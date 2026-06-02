"""Round-trip the Actuals tab back into inputs/actuals.json (Phase 7f, option A).

The CFO types weekly bank actuals into the workbook's "Actuals" tab, then runs:

    python -m src.ingest_actuals

which reads that tab and MERGES the entered weeks into inputs/actuals.json (the
machine source of truth the pipeline grades against). Run this BEFORE the next
pipeline refresh -- otherwise the regeneration re-seeds the Actuals tab from the
old JSON and the freshly-typed values are lost.

Only fully-entered weeks are ingested: a row must have a Week Start AND both
Actual Receipts and Actual Disbursements (Actual Ending Cash is optional). Rows
with a date but blank amounts are "not yet recorded" and skipped. Merge
semantics: entered weeks update/add; weeks already in the JSON but not on the
tab are preserved.

Fixed tab layout this reader depends on (see excel_writer._build_actuals_sheet):
row 1 header; rows 2+ are A=Week Start, B=Receipts, C=Disbursements, D=Ending Cash.
"""
from __future__ import annotations

import datetime as dt
import json
import logging
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from src.config import LOG_LEVEL, LOG_FORMAT, DATA_DIR
from src.output.excel_writer import OUTPUT_FILENAME, SHEET_ACTUALS
from src.calc.actuals_variance import INPUT_FILE as ACTUALS_JSON

logger = logging.getLogger(__name__)


def _coerce_week_key(v) -> Optional[str]:
    """Normalize a Week Start cell to an ISO date string, or None if unusable."""
    if v is None:
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return (v.date() if isinstance(v, dt.datetime) else v).isoformat()
    try:
        return dt.date.fromisoformat(str(v)[:10]).isoformat()
    except (TypeError, ValueError):
        return None


def read_actuals_tab(workbook_path: Path, sheet_name: str = SHEET_ACTUALS) -> dict:
    """Parse fully-entered rows of the Actuals tab into {week_iso: {...}}.

    A row counts only if it has a Week Start AND both receipts and disbursements.
    """
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    entered: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        week = _coerce_week_key(ws.cell(row=r, column=1).value)
        receipts = ws.cell(row=r, column=2).value
        disbursements = ws.cell(row=r, column=3).value
        ending_cash = ws.cell(row=r, column=4).value
        if week is None or receipts is None or disbursements is None:
            continue  # not a fully-recorded week
        rec = {"receipts": float(receipts), "disbursements": float(disbursements)}
        if ending_cash is not None:
            rec["ending_cash"] = float(ending_cash)
        entered[week] = rec
    return entered


def merge_into_json(entered: dict, json_path: Optional[Path] = None) -> tuple[int, int]:
    """Merge entered weeks into the actuals JSON. Returns (added, updated)."""
    json_path = Path(json_path) if json_path is not None else ACTUALS_JSON
    existing = {}
    if json_path.exists():
        with open(json_path) as f:
            existing = json.load(f) or {}
    added = updated = 0
    for week, rec in entered.items():
        if week in existing:
            if existing[week] != rec:
                updated += 1
        else:
            added += 1
        existing[week] = rec
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w") as f:
        json.dump(dict(sorted(existing.items())), f, indent=2)
        f.write("\n")
    return added, updated


def run(workbook_path: Optional[Path] = None) -> None:
    """Entrypoint: read the Actuals tab and merge entered weeks into the JSON."""
    logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
    workbook_path = Path(workbook_path) if workbook_path else DATA_DIR / OUTPUT_FILENAME
    if not workbook_path.exists():
        logger.warning("Workbook %s not found; nothing to ingest.", workbook_path)
        return

    entered = read_actuals_tab(workbook_path)
    if not entered:
        logger.info("No fully-entered actual weeks on the Actuals tab; %s unchanged.", ACTUALS_JSON)
        return
    added, updated = merge_into_json(entered)
    logger.info(
        "Ingested %d actual week(s) from %s -> %s (%d added, %d updated). "
        "Run the pipeline refresh next to grade them.",
        len(entered), workbook_path.name, ACTUALS_JSON, added, updated,
    )


if __name__ == "__main__":
    run()
