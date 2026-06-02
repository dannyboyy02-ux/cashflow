"""Render the 13-week cash forecast into a CFO-facing Excel workbook.

This is the CFO hand-off artifact: a single .xlsx that can be opened and read
without touching the pipeline. It reads the bucketed weekly tables from SQLite
plus the customer/vendor masters, and writes seven sheets (Phase 7d):

  1. "13-Week Forecast"    -- headline cash view: receipts, disbursements,
                              payroll, debt service, revolver, ending cash
  2. "Variance"            -- today vs prior snapshot (week level)
  3. "AR by Customer"      -- receipts (AR + open-SO) pivoted to customer x week
  4. "AP by Vendor"        -- disbursements (AP + open-PO) pivoted to vendor x week
  5. "Payroll"             -- weekly payroll drill-down (gross, burden, total)
  6. "Debt Service"        -- weekly per-loan principal + interest drill-down
  7. "Assumptions & Notes" -- as-of date, refresh stamp, methodology, caveats

The Forecast sheet's Payroll and Debt Service columns reference the matching
week row on their detail tabs, so the headline figures are transparently
traceable to their per-component build-up.

FORMULAS, NOT HARDCODED VALUES (per the xlsx skill standard):

The workbook is fully live. The only hardcoded numbers are the per-(entity,week)
cash figures on the AR/AP detail sheets (the model's source facts) and the
week-1 Beginning Cash input. Everything else is an Excel formula:

  - Forecast AR Receipts / AP Disbursements  -> cross-sheet =SUM links into the
    matching week column of the AR/AP detail sheets (green text: links).
  - Forecast Net Cash Flow                   -> =Receipts - Disbursements.
  - Forecast totals row                      -> =SUM down each column.
  - Detail-sheet Total column                -> =SUM across the 13 week columns.
  - Ending Cash = Beginning + Net; next week's Beginning = prior Ending, so the
    CFO's one starting-balance entry cascades through the whole horizon.

Color coding follows industry/skill convention: blue = hardcoded input (the
week-1 Beginning Cash, on a yellow "needs attention" fill), green = cross-sheet
link, black = same-sheet formula or source value.

Because openpyxl writes formulas as strings without evaluating them, run the
xlsx skill's scripts/recalc.py (LibreOffice) on the output to cache values and
scan for formula errors; Excel also recalculates them on open.
"""
from __future__ import annotations

import datetime as dt
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from src.config import (
    LOG_LEVEL,
    LOG_FORMAT,
    DATA_DIR,
    ONEDRIVE_DATA_PATH,
    FORECAST_HORIZON_WEEKS,
    ORG_NAME,
)
from src.db import get_connection
from src.calc.bucketing import (
    AR_OUTPUT_TABLE,
    AP_OUTPUT_TABLE,
    COMBINED_TABLE,
    COMBINED_DISBURSEMENTS_TABLE,
    SOURCE_AR,
    SOURCE_AP,
    monday_of_week,
)
from src.calc.po_payments_timing import (
    INVOICE_LAG_DAYS,
    OUTPUT_TABLE as PO_PAYMENTS_TABLE,
    SOURCE_STREAM_RBNI,
    SOURCE_STREAM_OUTSTANDING,
)
from src.calc.revolver import load_revolver_config
from src.calc.actuals_variance import load_actuals

logger = logging.getLogger(__name__)

OUTPUT_FILENAME = "cashflow_forecast.xlsx"

SHEET_FORECAST = "13-Week Forecast"
SHEET_VARIANCE = "Variance"
SHEET_AR = "AR by Customer"
SHEET_AP = "AP by Vendor"
SHEET_PAYROLL = "Payroll"
SHEET_DEBT = "Debt Service"
SHEET_REVOLVER = "Revolver"
SHEET_AVF = "Actual vs Forecast"
SHEET_ACTUALS = "Actuals"
SHEET_NOTES = "Assumptions & Notes"

ACTUALS_PLACEHOLDER = (
    "No closed weeks graded yet. Record weekly actuals on the Actuals tab, run "
    "`python -m src.ingest_actuals`, then refresh -- accuracy appears here."
)

# How many recently-closed weeks the Actuals input tab pre-lists for entry.
ACTUALS_LOOKBACK_WEEKS = 8

# Revolver-input seed values used to build the Assumptions "Revolver Inputs"
# block + named ranges when no live config is supplied (mirrors the keys in
# inputs/revolver_config.json).
REVOLVER_INPUT_DEFAULTS = {
    "facility_total": 60_000_000,
    "lc_carve_out": 2_545_000,
    "current_drawn_balance": 0,
    "beginning_cash": 1_196_696,
    "minimum_cash_target": 0,
    "sofr_rate": 0.0535,
    "spread": 0.0135,
}

# Revolver-tab column letters (1 row per forecast week). The Forecast sheet and
# the revolver-tab builder both reference these so the cross-sheet links agree.
REV_COL_BEGIN_CASH = "C"
REV_COL_INTEREST = "G"
REV_COL_DRAW = "K"
REV_COL_REPAY = "L"
REV_COL_ENDING_CASH = "N"

VARIANCE_PLACEHOLDER = (
    "No prior forecast available for comparison. "
    "Variance will appear after the next refresh."
)

# Currency, no decimals, $ prefix, negatives in red parentheses, zeros as a dash.
CURRENCY_FMT = "$#,##0_);[Red]($#,##0);-"
# Delta cells, favorable: positive green, negative red (parens), zero dash.
# Used for AR Δ and Net Δ, where an increase is good for cash.
DELTA_FMT = "[Green]$#,##0;[Red]($#,##0);-"
# Delta cells, unfavorable: colors flipped (positive red, negative green). Used
# for AP Δ, where an INCREASE in disbursements means more cash out -- bad.
DELTA_FMT_UNFAVORABLE = "[Red]$#,##0;[Green]($#,##0);-"
DATE_FMT = "yyyy-mm-dd"

# Professional, consistent font across the workbook (xlsx skill requirement).
FONT_NAME = "Arial"
F_BASE = Font(name=FONT_NAME)
F_HEADER = Font(name=FONT_NAME, bold=True)
F_TITLE = Font(name=FONT_NAME, bold=True, size=14)
F_INPUT = Font(name=FONT_NAME, color="0000FF")   # blue: hardcoded input
F_LINK = Font(name=FONT_NAME, color="008000")    # green: cross-sheet link

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")   # light gray
_INPUT_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")    # yellow: attention


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def load_table(name: str) -> pd.DataFrame:
    """Read an entire SQLite table into a DataFrame."""
    with get_connection() as conn:
        return pd.read_sql(f"SELECT * FROM {name}", conn)


def _load_optional(name: str) -> pd.DataFrame:
    """Read a table, or return an empty frame if it doesn't exist yet."""
    with get_connection() as conn:
        cur = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)
        )
        if cur.fetchone() is None:
            return pd.DataFrame()
        return pd.read_sql(f"SELECT * FROM {name}", conn)


def load_inputs() -> dict[str, pd.DataFrame]:
    """Load the source tables the writer needs.

    forecast_variance is optional: it doesn't exist until variance.run() has had
    two snapshots to compare, so it's read defensively (empty -> placeholder).
    """
    return {
        "receipts": load_table(AR_OUTPUT_TABLE),
        "disbursements": load_table(AP_OUTPUT_TABLE),
        "customers": load_table("bc_customers"),
        "vendors": load_table("bc_vendors"),
        "variance": _load_optional("forecast_variance"),
        "combined": _load_optional(COMBINED_TABLE),
        "combined_disbursements": _load_optional(COMBINED_DISBURSEMENTS_TABLE),
        "po_payments": _load_optional(PO_PAYMENTS_TABLE),
        "payroll": _load_optional("payroll_by_week"),
        "debt_principal": _load_optional("debt_principal_by_week"),
        "debt_interest": _load_optional("debt_interest_by_week"),
        "forecast_vs_actual": _load_optional("forecast_vs_actual"),
    }


def po_liabilities_diagnostics(po_payments: pd.DataFrame) -> dict:
    """Summarize the Phase 5 open-PO disbursements for the Assumptions sheet.

    Breaks RBNI and PO-outstanding totals down by Type, isolates the Item RBNI
    subtotal (the figure that reconciles to GL 21300), and reports the invoice
    lag. Returns zeros if po_payments is empty/None.
    """
    base = {
        "rbni_total": 0.0, "outstanding_total": 0.0, "phase5_total": 0.0,
        "rbni_by_type": {}, "outstanding_by_type": {}, "item_rbni_subtotal": 0.0,
        "invoice_lag_days": INVOICE_LAG_DAYS,
    }
    if po_payments is None or po_payments.empty or "source_stream" not in po_payments.columns:
        return base
    rbni = po_payments[po_payments["source_stream"] == SOURCE_STREAM_RBNI]
    out = po_payments[po_payments["source_stream"] == SOURCE_STREAM_OUTSTANDING]
    base["rbni_total"] = round(float(rbni["amount"].sum()), 2)
    base["outstanding_total"] = round(float(out["amount"].sum()), 2)
    base["phase5_total"] = round(base["rbni_total"] + base["outstanding_total"], 2)
    base["rbni_by_type"] = {k: round(float(v), 2) for k, v in rbni.groupby("Type")["amount"].sum().items()}
    base["outstanding_by_type"] = {k: round(float(v), 2) for k, v in out.groupby("Type")["amount"].sum().items()}
    base["item_rbni_subtotal"] = round(float(rbni[rbni["Type"] == "Item"]["amount"].sum()), 2)
    return base


def latest_refresh_timestamp(folder) -> Optional[dt.datetime]:
    """Return the mtime of the newest CSV in the OneDrive folder, or None.

    Used purely for the "source data refreshed" line on the notes sheet. Returns
    None (rather than raising) if the folder is missing or holds no CSVs, so the
    workbook still renders in environments without the live OneDrive mount.
    """
    folder = Path(folder)
    if not folder.exists():
        return None
    csvs = list(folder.glob("*.csv"))
    if not csvs:
        return None
    newest = max(csvs, key=lambda p: p.stat().st_mtime)
    return dt.datetime.fromtimestamp(newest.stat().st_mtime)


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------


def _style_header_row(ws: Worksheet, n_cols: int, row: int = 1) -> None:
    """Bold + light-gray fill across the first n_cols cells of a header row."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_HEADER
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _money(cell, font: Font = F_BASE) -> None:
    cell.number_format = CURRENCY_FMT
    cell.font = font


def _delta(cell, font: Font = F_BASE, fmt: str = DELTA_FMT) -> None:
    cell.number_format = fmt
    cell.font = font


def _coerce_date(v):
    """Best-effort convert a stored date string to a dt.date for an Excel cell."""
    if isinstance(v, (dt.date, dt.datetime)):
        return v
    try:
        return dt.date.fromisoformat(str(v)[:10])
    except (TypeError, ValueError):
        return v


def _wk_sums(df: pd.DataFrame, value_col: str) -> dict[int, float]:
    """Per-forecast-week sum of value_col as a {week: total} dict."""
    if df.empty or "forecast_week" not in df.columns:
        return {}
    return df.groupby("forecast_week")[value_col].sum().to_dict()


# ---------------------------------------------------------------------------
# Sheet 1 -- 13-Week Forecast
# ---------------------------------------------------------------------------


def _build_forecast_sheet(
    ws: Worksheet,
    as_of_date: dt.date,
    horizon_weeks: int,
    ar_data_rows: int,
    ap_data_rows: int,
    debt_total_col: str = "O",
) -> None:
    """Headline 11-column forecast summary (Phase 7e -- pure formula view):
    A  Forecast Week
    B  Week Start
    C  Beginning Cash          (='Revolver'!C{r}  -- revolver tab Begin Cash)
    D  AR + SO Receipts        (cross-sheet SUM from 'AR by Customer')
    E  AP + PO Disbursements   (cross-sheet SUM from 'AP by Vendor')
    F  Payroll                 (cross-sheet ref to 'Payroll' Total, same row)
    G  Debt Service            (cross-sheet ref to 'Debt Service' Total, same row)
    H  Revolver Net            (='Revolver'!K{r}-'Revolver'!L{r}  -- draw - repay)
    I  Revolver Interest       (='Revolver'!G{r})
    J  Net Cash Flow           (=D-E-F-G+H-I)
    K  Ending Cash             (='Revolver'!N{r}  -- revolver tab Ending Cash)

    EVERY cell is a formula -- there are no static revolver values anywhere on
    this sheet (Phase 7e). The revolver math is visible on the Revolver tab and
    referenced here; Beginning/Ending cash come straight from the Revolver tab,
    so clicking any of those cells traces into the plug logic. AR/AP use =SUM
    over all entity rows; Payroll/Debt reference their detail tabs same-row.
    """
    headers = [
        "Forecast Week", "Week Start", "Beginning Cash",
        "AR + SO Receipts", "AP + PO Disbursements",
        "Payroll", "Debt Service",
        "Revolver Net", "Revolver Interest",
        "Net Cash Flow", "Ending Cash",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    week_1_monday = monday_of_week(as_of_date)
    ar_last = 1 + ar_data_rows if ar_data_rows > 0 else 2
    ap_last = 1 + ap_data_rows if ap_data_rows > 0 else 2
    rev = f"'{SHEET_REVOLVER}'"

    for i in range(horizon_weeks):
        wk = i + 1
        r = i + 2  # worksheet row (1 = header); matches the Revolver-tab row.
        week_start = week_1_monday + dt.timedelta(days=7 * i)

        ws.cell(row=r, column=1, value=wk).font = F_BASE
        date_cell = ws.cell(row=r, column=2, value=week_start)
        date_cell.number_format = DATE_FMT
        date_cell.font = F_BASE

        # C: Beginning Cash — from the Revolver tab (its Begin Cash column).
        _money(ws.cell(row=r, column=3, value=f"={rev}!{REV_COL_BEGIN_CASH}{r}"), F_LINK)

        # D/E: AR + SO Receipts / AP + PO Disbursements — cross-sheet SUM (green).
        wk_col = get_column_letter(2 + wk)   # C = wk1, D = wk2 … O = wk13
        _money(ws.cell(row=r, column=4,
                       value=f"=SUM('{SHEET_AR}'!{wk_col}2:{wk_col}{ar_last})"), F_LINK)
        _money(ws.cell(row=r, column=5,
                       value=f"=SUM('{SHEET_AP}'!{wk_col}2:{wk_col}{ap_last})"), F_LINK)

        # F: Payroll — Total Payroll for this week (Payroll tab col F, same row).
        _money(ws.cell(row=r, column=6, value=f"='{SHEET_PAYROLL}'!F{r}"), F_LINK)

        # G: Debt Service — Total Debt Service for this week (same row).
        _money(ws.cell(row=r, column=7, value=f"='{SHEET_DEBT}'!{debt_total_col}{r}"), F_LINK)

        # H/I: Revolver Net (draw - repay) and Interest — from the Revolver tab.
        _money(ws.cell(row=r, column=8,
                       value=f"={rev}!{REV_COL_DRAW}{r}-{rev}!{REV_COL_REPAY}{r}"), F_LINK)
        _money(ws.cell(row=r, column=9, value=f"={rev}!{REV_COL_INTEREST}{r}"), F_LINK)

        # J: Net Cash Flow = Receipts - all outflows + Revolver Net - Rev Interest.
        _money(ws.cell(row=r, column=10,
                       value=f"=D{r}-E{r}-F{r}-G{r}+H{r}-I{r}"))
        # K: Ending Cash — from the Revolver tab (its Ending Cash column).
        _money(ws.cell(row=r, column=11, value=f"={rev}!{REV_COL_ENDING_CASH}{r}"), F_LINK)

    # Totals row (sums for all cash-movement columns; C and K not summed).
    tr = horizon_weeks + 2
    last = horizon_weeks + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = F_HEADER
    for col in (4, 5, 6, 7, 8, 9, 10):   # D through J
        letter = get_column_letter(col)
        cell = ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{last})")
        _money(cell, F_HEADER)

    ws.freeze_panes = "A2"
    _set_widths(ws, {
        "A": 14, "B": 12, "C": 16, "D": 18, "E": 20,
        "F": 14, "G": 14, "H": 14, "I": 18, "J": 16, "K": 16,
    })


# ---------------------------------------------------------------------------
# Sheets 2 & 3 -- entity x week pivots
# ---------------------------------------------------------------------------


def _build_entity_sheet(
    ws: Worksheet,
    df: pd.DataFrame,
    master: pd.DataFrame,
    entity_col: str,
    value_col: str,
    number_header: str,
    name_header: str,
    horizon_weeks: int,
) -> int:
    """One row per entity with any in-horizon cash, weeks across, sorted by Total.

    Week values are the model's source numbers; the Total column is a live
    =SUM across the 13 week columns. Returns the number of data rows written.
    """
    headers = (
        [number_header, name_header]
        + [f"Week {w}" for w in range(1, horizon_weeks + 1)]
        + ["Total"]
    )
    ws.append(headers)
    _style_header_row(ws, len(headers))

    name_map: dict = {}
    if not master.empty and "number" in master.columns and "displayName" in master.columns:
        name_map = dict(zip(master["number"], master["displayName"]))

    n_rows = 0
    first_week_col = 3                       # column C
    last_week_col = 2 + horizon_weeks        # column O for 13 weeks
    total_col = last_week_col + 1            # column P
    first_letter = get_column_letter(first_week_col)
    last_letter = get_column_letter(last_week_col)

    if not df.empty and "forecast_week" in df.columns:
        pivot = df.pivot_table(
            index=entity_col,
            columns="forecast_week",
            values=value_col,
            aggfunc="sum",
            fill_value=0.0,
        )
        built = []
        for ent in pivot.index:
            week_vals = [
                float(pivot.loc[ent, w]) if w in pivot.columns else 0.0
                for w in range(1, horizon_weeks + 1)
            ]
            built.append((ent, name_map.get(ent, ""), week_vals, sum(week_vals)))
        built.sort(key=lambda x: x[3], reverse=True)

        for ent, name, week_vals, _total in built:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=ent).font = F_BASE
            ws.cell(row=r, column=2, value=name).font = F_BASE
            for j, v in enumerate(week_vals):
                _money(ws.cell(row=r, column=first_week_col + j, value=v))
            # Total column: live SUM across the week cells.
            _money(
                ws.cell(
                    row=r,
                    column=total_col,
                    value=f"=SUM({first_letter}{r}:{last_letter}{r})",
                )
            )
            n_rows += 1

    ws.freeze_panes = "C2"  # keep number + name + header visible while scrolling
    widths = {"A": 16, "B": 32}
    for w in range(1, horizon_weeks + 1):
        widths[get_column_letter(2 + w)] = 12
    widths[get_column_letter(total_col)] = 16
    _set_widths(ws, widths)
    return n_rows


# ---------------------------------------------------------------------------
# Payroll detail tab (vertical: weeks as rows, components as columns)
# ---------------------------------------------------------------------------


def _build_payroll_sheet(ws: Worksheet, payroll: Optional[pd.DataFrame], horizon_weeks: int) -> None:
    """Payroll drill-down: one row per week showing gross wages, burden, total.

    Week columns on this sheet are NOT the week-across format used by AR/AP
    entity sheets; instead weeks are rows (13 data rows). The Forecast sheet
    references the Total Payroll column (F) by row number.

    Row layout (same week numbering as the Forecast sheet — row r = week r-1):
      Row 1  : Header
      Row 2  : Week 1 data  (r = wk + 1)
      ...
      Row 14 : Week 13 data
      Row 15 : Totals
    """
    headers = [
        "Forecast Week", "Week Start",
        "Gross Wages", "Employer Burden %", "Employer Burden ($)", "Total Payroll",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    by_week: dict[int, dict] = {}
    if payroll is not None and not payroll.empty:
        for _, row in payroll.iterrows():
            by_week[int(row["forecast_week"])] = row.to_dict()

    tot_gross = tot_burden = tot_total = 0.0
    for i in range(horizon_weeks):
        wk = i + 1
        r = i + 2
        row = by_week.get(wk, {})
        gross = float(row.get("gross_wages", 0.0))
        pct = float(row.get("employer_burden_pct", 0.0))
        burden_amt = round(gross * pct, 2)
        total = float(row.get("total_payroll", 0.0))
        ws.cell(row=r, column=1, value=wk).font = F_BASE
        ws_date = row.get("week_start_date")
        if ws_date:
            dc = ws.cell(row=r, column=2, value=_coerce_date(ws_date))
            dc.number_format = DATE_FMT
            dc.font = F_BASE
        _money(ws.cell(row=r, column=3, value=gross))
        pct_cell = ws.cell(row=r, column=4, value=pct)
        pct_cell.number_format = "0.00%"
        pct_cell.font = F_BASE
        _money(ws.cell(row=r, column=5, value=burden_amt))
        _money(ws.cell(row=r, column=6, value=total))   # ← Forecast col F refs this row
        tot_gross += gross
        tot_burden += burden_amt
        tot_total += total

    # Totals row.
    tr = horizon_weeks + 2
    ws.cell(row=tr, column=1, value="TOTAL").font = F_HEADER
    for col, val in ((3, tot_gross), (5, tot_burden), (6, tot_total)):
        cell = ws.cell(row=tr, column=col, value=round(val, 2))
        _money(cell, F_HEADER)

    ws.freeze_panes = "A2"
    _set_widths(ws, {"A": 14, "B": 12, "C": 16, "D": 18, "E": 18, "F": 16})


# ---------------------------------------------------------------------------
# Debt Service detail tab (vertical: weeks as rows, loans as columns)
# ---------------------------------------------------------------------------


def _build_debt_service_sheet(
    ws: Worksheet,
    debt_principal: Optional[pd.DataFrame],
    debt_interest: Optional[pd.DataFrame],
    horizon_weeks: int,
) -> str:
    """Debt service drill-down showing per-loan principal, per-loan interest,
    and a Total Debt Service column (referenced by the Forecast sheet). Returns
    the column letter of the Total Debt Service column (depends on loan count;
    column O for the standard 5 loans).

    Column layout (per-week rows; row r = wk + 1 matching Payroll tab):
      A  Forecast Week   B  Week Start
      C-G  Loan 1-5 Principal    H  Total Principal
      I-M  Loan 1-5 Interest     N  Total Interest
      O  Total Debt Service   ← Forecast col G references this
    """
    # Gather sorted loan names from whichever table has data.
    loan_names: list[str] = []
    for df in (debt_principal, debt_interest):
        if df is not None and not df.empty and "loan_name" in df.columns:
            for ln in df["loan_name"].dropna().unique():
                if ln not in loan_names:
                    loan_names.append(ln)
    loan_names = sorted(loan_names)
    n_loans = len(loan_names) if loan_names else 0

    # Build {week -> {loan_name -> principal/interest}}
    def _by_week_loan(df, val_col):
        result: dict[int, dict[str, float]] = {wk: {} for wk in range(1, horizon_weeks + 1)}
        if df is None or df.empty:
            return result
        for _, row in df.iterrows():
            wk = int(row["forecast_week"])
            if 1 <= wk <= horizon_weeks:
                result[wk][str(row["loan_name"])] = float(row[val_col])
        return result

    prin_data = _by_week_loan(debt_principal, "principal")
    int_data = _by_week_loan(debt_interest, "interest")

    # Build headers (dynamic on loan count, but expect 5 loans).
    p_cols = [f"{ln} Principal" for ln in loan_names] if loan_names else ["Principal"]
    i_cols = [f"{ln} Interest" for ln in loan_names] if loan_names else ["Interest"]
    headers = (["Forecast Week", "Week Start"]
               + p_cols + ["Total Principal"]
               + i_cols + ["Total Interest", "Total Debt Service"])
    ws.append(headers)
    _style_header_row(ws, len(headers))

    n_p = max(n_loans, 1)   # columns for principal
    n_i = max(n_loans, 1)   # columns for interest
    # column index offsets (1-based):
    #  A=1, B=2, C=3 .. C+n_p-1 = principal per loan
    #  C+n_p = Total Principal
    #  C+n_p+1 .. C+n_p+n_i = interest per loan
    #  C+n_p+n_i+1 = Total Interest
    #  C+n_p+n_i+2 = Total Debt Service  ← col O for 5 loans (= col 16)
    tp_col = 2 + n_p + 1       # Total Principal
    ti_col = tp_col + n_i + 1  # Total Interest
    tds_col = ti_col + 1       # Total Debt Service (Forecast col G references this)

    tot_principal = tot_interest = 0.0
    for i in range(horizon_weeks):
        wk = i + 1
        r = i + 2
        ws.cell(row=r, column=1, value=wk).font = F_BASE

        # Grab week_start_date from whichever table has it.
        wsd = None
        for df in (debt_principal, debt_interest):
            if df is not None and not df.empty:
                rows_wk = df[df["forecast_week"] == wk]
                if not rows_wk.empty:
                    wsd = rows_wk.iloc[0]["week_start_date"]
                    break
        if wsd:
            dc = ws.cell(row=r, column=2, value=_coerce_date(wsd))
            dc.number_format = DATE_FMT
            dc.font = F_BASE

        # Per-loan principal.
        week_p = sum_p = 0.0
        for j, ln in enumerate(loan_names):
            v = prin_data[wk].get(ln, 0.0)
            _money(ws.cell(row=r, column=3 + j, value=v))
            sum_p += v
        _money(ws.cell(row=r, column=tp_col, value=round(sum_p, 2)), F_HEADER if sum_p else F_BASE)

        # Per-loan interest.
        sum_i = 0.0
        for j, ln in enumerate(loan_names):
            v = int_data[wk].get(ln, 0.0)
            _money(ws.cell(row=r, column=tp_col + 1 + j, value=v))
            sum_i += v
        _money(ws.cell(row=r, column=ti_col, value=round(sum_i, 2)), F_HEADER if sum_i else F_BASE)

        # Total Debt Service  ← Forecast sheet references this column.
        tds = round(sum_p + sum_i, 2)
        _money(ws.cell(row=r, column=tds_col, value=tds), F_HEADER if tds else F_BASE)
        tot_principal += sum_p
        tot_interest += sum_i

    # Totals row.
    tr = horizon_weeks + 2
    ws.cell(row=tr, column=1, value="TOTAL").font = F_HEADER
    _money(ws.cell(row=tr, column=tp_col, value=round(tot_principal, 2)), F_HEADER)
    _money(ws.cell(row=tr, column=ti_col, value=round(tot_interest, 2)), F_HEADER)
    _money(ws.cell(row=tr, column=tds_col, value=round(tot_principal + tot_interest, 2)), F_HEADER)

    ws.freeze_panes = "C2"
    _set_widths(ws, {"A": 12, "B": 12,
                     **{get_column_letter(3 + j): 14 for j in range(n_p + n_i + 3)}})
    return get_column_letter(tds_col)


# ---------------------------------------------------------------------------
# AR by Customer -- open-AR + open-SO contributions, one row per (customer, source)
# ---------------------------------------------------------------------------


def _build_ar_by_customer_sheet(
    ws: Worksheet,
    ar_long: pd.DataFrame,
    master: pd.DataFrame,
    horizon_weeks: int,
) -> int:
    """AR-by-customer detail with a Source column distinguishing open-AR vs open-SO.

    Accepts either combined_receipts_by_week (has a "source" column) or
    ar_receipts_by_week (no source -> treated as all open_ar). One row per
    (customer, source), sorted by row total descending. The week columns stay at
    C..O and Total at P -- identical to the AP sheet -- so the forecast sheet's
    cross-sheet SUM over column C..O captures BOTH sources automatically; Source
    is an extra trailing column (Q). Returns the number of data rows written.
    """
    headers = (
        ["Customer Number", "Customer Name"]
        + [f"Week {w}" for w in range(1, horizon_weeks + 1)]
        + ["Total", "Source"]
    )
    ws.append(headers)
    _style_header_row(ws, len(headers))

    name_map: dict = {}
    if not master.empty and "number" in master.columns and "displayName" in master.columns:
        name_map = dict(zip(master["number"], master["displayName"]))

    first_week_col = 3
    last_week_col = 2 + horizon_weeks         # O for 13 weeks
    total_col = last_week_col + 1             # P
    source_col = total_col + 1                # Q
    first_letter = get_column_letter(first_week_col)
    last_letter = get_column_letter(last_week_col)

    n_rows = 0
    if ar_long is not None and not ar_long.empty and "forecast_week" in ar_long.columns:
        df = ar_long.copy()
        if "source" not in df.columns:
            df["source"] = SOURCE_AR
        built = []
        for (cust, source), g in df.groupby(["customerNumber", "source"]):
            wk_sum = g.groupby("forecast_week")["receipts"].sum()
            week_vals = [float(wk_sum.get(w, 0.0)) for w in range(1, horizon_weeks + 1)]
            built.append((cust, name_map.get(cust, ""), source, week_vals, sum(week_vals)))
        built.sort(key=lambda x: x[4], reverse=True)

        for cust, name, source, week_vals, _total in built:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=cust).font = F_BASE
            ws.cell(row=r, column=2, value=name).font = F_BASE
            for j, v in enumerate(week_vals):
                _money(ws.cell(row=r, column=first_week_col + j, value=v))
            _money(ws.cell(row=r, column=total_col,
                           value=f"=SUM({first_letter}{r}:{last_letter}{r})"))
            ws.cell(row=r, column=source_col, value=source).font = F_BASE
            n_rows += 1

    ws.freeze_panes = "C2"
    widths = {"A": 16, "B": 32}
    for w in range(1, horizon_weeks + 1):
        widths[get_column_letter(2 + w)] = 12
    widths[get_column_letter(total_col)] = 16
    widths[get_column_letter(source_col)] = 10
    _set_widths(ws, widths)
    return n_rows


# ---------------------------------------------------------------------------
# AP by Vendor -- open-AP + open-PO (RBNI / outstanding), one row per (vendor, source)
# ---------------------------------------------------------------------------


def _build_ap_by_vendor_sheet(
    ws: Worksheet,
    ap_long: pd.DataFrame,
    master: pd.DataFrame,
    horizon_weeks: int,
) -> int:
    """AP-by-vendor detail with a Source column (open_ap / po_rbni / po_outstanding).

    Mirror of _build_ar_by_customer_sheet on the AP side. Accepts either
    combined_disbursements_by_week (has a "source" column) or
    ap_disbursements_by_week (no source -> treated as all open_ap). One row per
    (vendor, source), sorted by row total descending. Week columns stay at C..O
    and Total at P so the forecast sheet's cross-sheet SUM over C..O captures all
    sources; Source is the trailing column (Q). Returns the data-row count.
    """
    headers = (
        ["Vendor Number", "Vendor Name"]
        + [f"Week {w}" for w in range(1, horizon_weeks + 1)]
        + ["Total", "Source"]
    )
    ws.append(headers)
    _style_header_row(ws, len(headers))

    name_map: dict = {}
    if not master.empty and "number" in master.columns and "displayName" in master.columns:
        name_map = dict(zip(master["number"], master["displayName"]))

    first_week_col = 3
    last_week_col = 2 + horizon_weeks         # O for 13 weeks
    total_col = last_week_col + 1             # P
    source_col = total_col + 1                # Q
    first_letter = get_column_letter(first_week_col)
    last_letter = get_column_letter(last_week_col)

    n_rows = 0
    if ap_long is not None and not ap_long.empty and "forecast_week" in ap_long.columns:
        df = ap_long.copy()
        if "source" not in df.columns:
            df["source"] = SOURCE_AP
        built = []
        for (vendor, source), g in df.groupby(["vendorNumber", "source"]):
            wk_sum = g.groupby("forecast_week")["disbursements"].sum()
            week_vals = [float(wk_sum.get(w, 0.0)) for w in range(1, horizon_weeks + 1)]
            built.append((vendor, name_map.get(vendor, ""), source, week_vals, sum(week_vals)))
        built.sort(key=lambda x: x[4], reverse=True)

        for vendor, name, source, week_vals, _total in built:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=vendor).font = F_BASE
            ws.cell(row=r, column=2, value=name).font = F_BASE
            for j, v in enumerate(week_vals):
                _money(ws.cell(row=r, column=first_week_col + j, value=v))
            _money(ws.cell(row=r, column=total_col,
                           value=f"=SUM({first_letter}{r}:{last_letter}{r})"))
            ws.cell(row=r, column=source_col, value=source).font = F_BASE
            n_rows += 1

    ws.freeze_panes = "C2"
    widths = {"A": 16, "B": 32}
    for w in range(1, horizon_weeks + 1):
        widths[get_column_letter(2 + w)] = 12
    widths[get_column_letter(total_col)] = 16
    widths[get_column_letter(source_col)] = 10
    _set_widths(ws, widths)
    return n_rows


# ---------------------------------------------------------------------------
# Revolver tab -- the plug, as VISIBLE Excel math (Phase 7e)
# ---------------------------------------------------------------------------


def _build_revolver_sheet(ws: Worksheet, as_of_date: dt.date, horizon_weeks: int) -> None:
    """Sequential revolver plug rendered entirely in Excel formulas.

    One row per forecast week (row r = week r-1, matching the Forecast and
    detail tabs). Every computed cell is a formula referencing named-range
    inputs (from the Assumptions Revolver Inputs block), the Forecast sheet
    (Inflows / Base Outflows), and the prior row (sequential carry). Click any
    cell to trace the math -- no static values.

    Columns: A Forecast Week | B Week Start | C Begin Cash | D Inflows |
    E Base Outflows | F Begin Revolver Balance | G Interest Accrued |
    H Pre-Revolver Cash | I Min Cash Target | J Available Capacity | K Draw |
    L Repay | M Ending Revolver Balance | N Ending Cash | O Capacity Breached.

    Interest is ROUND()ed to the cent so the Excel result matches revolver.py
    (which rounds interest the same way) to the penny.
    """
    headers = [
        "Forecast Week", "Week Start", "Begin Cash", "Inflows", "Base Outflows",
        "Begin Revolver Balance", "Interest Accrued", "Pre-Revolver Cash",
        "Min Cash Target", "Available Capacity", "Draw", "Repay",
        "Ending Revolver Balance", "Ending Cash", "Capacity Breached",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    week_1_monday = monday_of_week(as_of_date)
    fc = f"'{SHEET_FORECAST}'"

    for i in range(horizon_weeks):
        wk = i + 1
        r = i + 2
        week_start = week_1_monday + dt.timedelta(days=7 * i)

        ws.cell(row=r, column=1, value=wk).font = F_BASE
        dcell = ws.cell(row=r, column=2, value=week_start)
        dcell.number_format = DATE_FMT
        dcell.font = F_BASE

        # C Begin Cash: wk1 anchors to the input; later weeks chain off prior N.
        c_begin = "=Beginning_Cash_Wk1" if wk == 1 else f"=N{r - 1}"
        _money(ws.cell(row=r, column=3, value=c_begin))
        # D Inflows / E Base Outflows: pulled from the Forecast summary row.
        _money(ws.cell(row=r, column=4, value=f"={fc}!D{r}"), F_LINK)
        _money(ws.cell(row=r, column=5, value=f"={fc}!E{r}+{fc}!F{r}+{fc}!G{r}"), F_LINK)
        # F Begin Revolver: wk1 anchors to currently-drawn; later weeks chain off prior M.
        f_begin = "=Currently_Drawn" if wk == 1 else f"=M{r - 1}"
        _money(ws.cell(row=r, column=6, value=f_begin))
        # G Interest Accrued (on opening balance), rounded to the cent.
        _money(ws.cell(row=r, column=7, value=f"=ROUND(F{r}*Annual_Rate/52,2)"))
        # H Pre-Revolver Cash.
        _money(ws.cell(row=r, column=8, value=f"=C{r}+D{r}-E{r}-G{r}"))
        # I Min Cash Target (per-row, for visibility).
        _money(ws.cell(row=r, column=9, value="=Min_Cash_Target"))
        # J Available Capacity.
        _money(ws.cell(row=r, column=10, value=f"=Max_Capacity-F{r}"))
        # K Draw: cover the shortfall to target, bounded by capacity.
        _money(ws.cell(row=r, column=11,
                       value=f"=IF(H{r}<I{r},MIN(I{r}-H{r},J{r}),0)"))
        # L Repay: sweep excess above target, bounded by the outstanding balance.
        _money(ws.cell(row=r, column=12,
                       value=f"=IF(AND(H{r}>I{r},F{r}>0),MIN(H{r}-I{r},F{r}),0)"))
        # M Ending Revolver Balance.
        _money(ws.cell(row=r, column=13, value=f"=F{r}+K{r}-L{r}"))
        # N Ending Cash.
        _money(ws.cell(row=r, column=14, value=f"=H{r}+K{r}-L{r}"))
        # O Capacity Breached flag.
        breach = f'=IF(AND(H{r}<I{r},I{r}-H{r}>J{r}),"BREACHED","")'
        bc = ws.cell(row=r, column=15, value=breach)
        bc.font = F_BASE

    ws.freeze_panes = "C2"
    widths = {"A": 13, "B": 12}
    for c in range(3, 15):
        widths[get_column_letter(c)] = 16
    widths["O"] = 16
    _set_widths(ws, widths)


# ---------------------------------------------------------------------------
# Variance sheet -- today vs prior snapshot (week level)
# ---------------------------------------------------------------------------


def _build_variance_sheet(
    ws: Worksheet,
    variance: Optional[pd.DataFrame],
    horizon_weeks: int,
) -> None:
    """Week-level day-over-day variance. AR/AP today/prior are values; deltas and
    net rows are formulas. Delta columns use the green/red delta format.

    When there is no prior snapshot (variance empty/None), writes a single
    placeholder cell and leaves the rest of the sheet blank.
    """
    if variance is None or len(variance) == 0:
        cell = ws.cell(row=1, column=1, value=VARIANCE_PLACEHOLDER)
        cell.font = F_BASE
        ws.column_dimensions["A"].width = 80
        return

    headers = [
        "Forecast Week", "Week Start",
        "AR Today", "AR Prior", "AR Δ",
        "AP Today", "AP Prior", "AP Δ",
        "Net Today", "Net Prior", "Net Δ",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    by_week = {int(r["forecast_week"]): r for _, r in variance.iterrows()}

    for i in range(horizon_weeks):
        wk = i + 1
        r = i + 2
        row = by_week.get(wk)

        ws.cell(row=r, column=1, value=wk).font = F_BASE
        if row is not None and row["week_start_date"] is not None:
            dcell = ws.cell(row=r, column=2, value=_coerce_date(row["week_start_date"]))
            dcell.number_format = DATE_FMT
            dcell.font = F_BASE

        art = float(row["ar_receipts_today"]) if row is not None else 0.0
        arp = float(row["ar_receipts_prior"]) if row is not None else 0.0
        apt = float(row["ap_disbursements_today"]) if row is not None else 0.0
        app = float(row["ap_disbursements_prior"]) if row is not None else 0.0

        _money(ws.cell(row=r, column=3, value=art))   # AR Today
        _money(ws.cell(row=r, column=4, value=arp))   # AR Prior
        _delta(ws.cell(row=r, column=5, value=f"=C{r}-D{r}"))   # AR delta
        _money(ws.cell(row=r, column=6, value=apt))   # AP Today
        _money(ws.cell(row=r, column=7, value=app))   # AP Prior
        # AP delta uses the unfavorable format: a rise in disbursements is bad.
        _delta(ws.cell(row=r, column=8, value=f"=F{r}-G{r}"), fmt=DELTA_FMT_UNFAVORABLE)
        _money(ws.cell(row=r, column=9, value=f"=C{r}-F{r}"))   # Net Today
        _money(ws.cell(row=r, column=10, value=f"=D{r}-G{r}"))  # Net Prior
        _delta(ws.cell(row=r, column=11, value=f"=I{r}-J{r}"))  # Net delta

    tr = horizon_weeks + 2
    last_data_row = horizon_weeks + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = F_HEADER
    for col in range(3, 12):
        letter = get_column_letter(col)
        cell = ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{last_data_row})")
        if col == 8:  # AP Δ total: unfavorable coloring
            _delta(cell, F_HEADER, fmt=DELTA_FMT_UNFAVORABLE)
        elif col in (5, 11):  # AR Δ / Net Δ totals: favorable coloring
            _delta(cell, F_HEADER)
        else:
            _money(cell, F_HEADER)

    ws.freeze_panes = "C2"
    widths = {"A": 14, "B": 12}
    for c in range(3, 12):
        widths[get_column_letter(c)] = 14
    _set_widths(ws, widths)


# ---------------------------------------------------------------------------
# Actual vs Forecast -- weekly forecast-accuracy scorecard (display)
# ---------------------------------------------------------------------------


def _build_actual_vs_forecast_sheet(
    ws: Worksheet,
    fva: Optional[pd.DataFrame],
) -> None:
    """Render forecast_vs_actual: one row per graded (closed) week.

    receipts/net variances use the favorable delta format (actual > forecast is
    good for cash, green); the disbursements variance uses the unfavorable format
    (actual > forecast means we paid more than expected, red). Placeholder when
    nothing has been graded yet.
    """
    if fva is None or fva.empty:
        cell = ws.cell(row=1, column=1, value=ACTUALS_PLACEHOLDER)
        cell.font = F_BASE
        ws.column_dimensions["A"].width = 90
        return

    headers = [
        "Week Start", "Forecast Receipts", "Actual Receipts", "Receipts Var",
        "Forecast Disb", "Actual Disb", "Disb Var",
        "Forecast Net", "Actual Net", "Net Var",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    src_cols = [
        "forecast_receipts", "actual_receipts", "receipts_variance",
        "forecast_disbursements", "actual_disbursements", "disbursements_variance",
        "forecast_net", "actual_net", "net_variance",
    ]
    fva = fva.sort_values("week_start_date")
    for _, row in fva.iterrows():
        r = ws.max_row + 1
        dcell = ws.cell(row=r, column=1, value=_coerce_date(row["week_start_date"]))
        dcell.number_format = DATE_FMT
        dcell.font = F_BASE
        for j, col in enumerate(src_cols):
            c = ws.cell(row=r, column=2 + j, value=float(row[col]))
            if col == "receipts_variance" or col == "net_variance":
                _delta(c)                                   # favorable: green +
            elif col == "disbursements_variance":
                _delta(c, fmt=DELTA_FMT_UNFAVORABLE)        # unfavorable: red +
            else:
                _money(c)

    # Totals row.
    tr = ws.max_row + 1
    last = tr - 1
    ws.cell(row=tr, column=1, value="TOTAL").font = F_HEADER
    for col in range(2, 11):
        letter = get_column_letter(col)
        cell = ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{last})")
        if col in (4, 10):       # receipts var, net var
            _delta(cell, F_HEADER)
        elif col == 7:           # disb var
            _delta(cell, F_HEADER, fmt=DELTA_FMT_UNFAVORABLE)
        else:
            _money(cell, F_HEADER)

    ws.freeze_panes = "B2"
    widths = {"A": 12}
    for c in range(2, 11):
        widths[get_column_letter(c)] = 15
    _set_widths(ws, widths)


# ---------------------------------------------------------------------------
# Actuals -- weekly bank actuals INPUT tab (round-trip via ingest_actuals)
# ---------------------------------------------------------------------------


def _build_actuals_sheet(
    ws: Worksheet,
    actuals: Optional[dict],
    as_of_date: dt.date,
    horizon_weeks: int,
) -> None:
    """Input tab where the CFO types actuals for closed weeks (yellow cells).

    Pre-lists the recently-closed weeks (ACTUALS_LOOKBACK_WEEKS before week 1)
    plus any weeks already recorded in inputs/actuals.json, seeding entered
    values. After typing, run `python -m src.ingest_actuals` to write these back
    to actuals.json before the next refresh (the JSON is the source of truth; this
    tab is re-seeded from it on each generation).

    Fixed layout for the ingest reader: row 1 header; rows 2+ are
    A=Week Start, B=Actual Receipts, C=Actual Disbursements, D=Actual Ending Cash.
    """
    actuals = actuals or {}
    headers = ["Week Start", "Actual Receipts", "Actual Disbursements", "Actual Ending Cash"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    note = ws.cell(row=1, column=6,
                   value="Enter actuals for CLOSED weeks (receipts + disbursements "
                         "both required), then run: python -m src.ingest_actuals  "
                         "before the next refresh.")
    note.font = F_BASE

    week_1_monday = monday_of_week(as_of_date)
    recent = {(week_1_monday - dt.timedelta(days=7 * k)).isoformat()
              for k in range(1, ACTUALS_LOOKBACK_WEEKS + 1)}
    weeks = sorted(recent | set(actuals.keys()))

    for i, wk in enumerate(weeks):
        r = i + 2
        rec = actuals.get(wk, {})
        dcell = ws.cell(row=r, column=1, value=_coerce_date(wk))
        dcell.number_format = DATE_FMT
        dcell.font = F_BASE
        for col, key in ((2, "receipts"), (3, "disbursements"), (4, "ending_cash")):
            cell = ws.cell(row=r, column=col, value=rec.get(key))   # None if unrecorded
            cell.fill = _INPUT_FILL
            _money(cell, F_INPUT)

    ws.freeze_panes = "B2"
    _set_widths(ws, {"A": 12, "B": 18, "C": 20, "D": 18})


# ---------------------------------------------------------------------------
# Sheet 4 -- Assumptions & Notes
# ---------------------------------------------------------------------------


def _build_assumptions_sheet(
    ws: Worksheet,
    as_of_date: dt.date,
    refresh_ts: Optional[dt.datetime],
    po_diagnostics: Optional[dict] = None,
    revolver_config: Optional[dict] = None,
) -> None:
    refresh_str = refresh_ts.isoformat(sep=" ", timespec="seconds") if refresh_ts else "unknown"
    lines: list[tuple[str, bool]] = [
        (f"{ORG_NAME} -- 13-Week Cash Flow Forecast", True),
        ("", False),
        (f"As-of date: {as_of_date.isoformat()}", False),
        (f"Source data refreshed: {refresh_str}", False),
        ("", False),
        ("Methodology", True),
        (
            "Receipts and disbursements are timed onto a Monday-anchored 13-week "
            "calendar grid (week 1 = the week containing the as-of date).",
            False,
        ),
        (
            "AR receipts: each open invoice is expected on postingDate + the "
            "customer's empirical DSO (days sales outstanding from trailing "
            "12-month sales). AP disbursements use two streams: (A) payments AP "
            "has already scheduled in Business Central with a future date, taken "
            "verbatim; and (B) open invoices estimated at postingDate + the "
            "vendor's empirical DPO (days payable outstanding).",
            False,
        ),
        (
            "Overdue clamp: any item whose estimated date is already past is "
            "pulled forward to the as-of date (week 1), on the assumption it "
            "collects/pays imminently. Such rows are flagged was_overdue upstream.",
            False,
        ),
        ("", False),
        ("Timing methods (timing_method column upstream)", True),
        ("  ratio          -- empirical DSO/DPO from trailing 12-month activity", False),
        ("  terms_fallback -- contractual payment terms (no trailing activity to measure)", False),
        ("  no_balance     -- master balance <= 0; treated as immediate, clamped to as-of", False),
        ("  master_fallback-- entity missing from the DSO/DPO table; used the row's due-days", False),
        ("  scheduled      -- AP payment pre-loaded in Business Central (Stream A), date as-is", False),
        ("", False),
        ("Caveats", True),
        (
            "This is an automated forecast from current open AR/AP, open SO/PO, "
            "payroll, debt service, and a revolver plug. It does NOT yet include:",
            False,
        ),
        ("  - capital expenditures", False),
        ("  - tax payments", False),
        ("  - new billings beyond current open AR / new POs beyond current open PO", False),
        ("", False),
        ("Revolver math is visible on the Revolver tab (every cell a formula).", False),
        ("Cash & revolver assumptions are the yellow input cells in the Revolver", False),
        ("Inputs block below; edit them and the model recomputes on open.", False),
    ]

    # Phase 5 / PO Liabilities Diagnostic block. RBNI (received-but-not-invoiced)
    # and PO outstanding are now INCLUDED in the AP forecast (Phase 5), so they
    # are no longer in the caveat list above.
    diag = po_diagnostics or po_liabilities_diagnostics(None)
    lines += [
        ("", False),
        ("Phase 5 / PO Liabilities Diagnostic", True),
        (f"  Invoice lag (goods received -> vendor invoice posted): {diag['invoice_lag_days']} days", False),
        (f"  Total RBNI extracted (received, not yet invoiced): ${diag['rbni_total']:,.0f}", False),
    ]
    for t, v in sorted(diag["rbni_by_type"].items(), key=lambda kv: -kv[1]):
        lines.append((f"      RBNI by Type -- {t}: ${v:,.0f}", False))
    lines.append((f"  Total PO outstanding extracted (ordered, not yet received): ${diag['outstanding_total']:,.0f}", False))
    for t, v in sorted(diag["outstanding_by_type"].items(), key=lambda kv: -kv[1]):
        lines.append((f"      PO outstanding by Type -- {t}: ${v:,.0f}", False))
    lines += [
        (f"  Total Phase 5 future AP cash added (RBNI + outstanding): ${diag['phase5_total']:,.0f}", False),
        (f"  Item-only RBNI subtotal: ${diag['item_rbni_subtotal']:,.0f}", False),
        (
            "  GL 21300 reconciliation: Item RBNI subtotal should approximately "
            "match the GL 21300 (Invt. Accrual Acc. Interim) balance net of "
            "pending purchase credit memos. Manual check; PO credit memo handling "
            "is a queued v2 enhancement.",
            False,
        ),
        ("", False),
        ("Disbursement certainty tiers (FP&A convention)", True),
        (
            "  Tier 1 - contractual / committed (high certainty): open AP invoices, "
            "scheduled BC payments, debt service, payroll. Due on known terms.",
            False,
        ),
        (
            "  Tier 2 - received, not yet invoiced (RBNI): goods are in hand and a "
            "vendor invoice is imminent (~7 days), so conversion to cash is near-"
            "certain; timing is the only estimate.",
            False,
        ),
        (
            "  Tier 3 - ordered, not yet received (PO outstanding): committed but "
            "LOWEST in-horizon certainty -- the order can still be changed, "
            "delayed, or cancelled before receipt. Included here at 100% (gross), "
            "which is a deliberately CONSERVATIVE liquidity view (overstates near-"
            "term outflow). It is the single largest swing factor in the revolver "
            "draw. A probability/conversion haircut on Tier 3 is the recommended "
            "next refinement (config-driven, e.g. po_outstanding_haircut_pct).",
            False,
        ),
        (
            "  Read the headline with this in mind: peak revolver draw is inflated "
            "by Tier 3 at full weight. The Revolver tab's Ending Revolver Balance "
            "is the point-in-time draw, well inside the facility once Tier 3 is "
            "weighted to realistic conversion.",
            False,
        ),
    ]

    for i, (text, is_header) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if i == 1:
            cell.font = F_TITLE
        elif is_header:
            cell.font = F_HEADER
        else:
            cell.font = F_BASE

    # --- Revolver Inputs block: yellow input cells + workbook named ranges. ---
    # Named ranges let downstream formulas read =Begin_Bal*Annual_Rate/52 rather
    # than =Assumptions!$B$NN (the FP&A-standard pattern). Seeded from the live
    # revolver config when supplied, else from REVOLVER_INPUT_DEFAULTS.
    cfg = {**REVOLVER_INPUT_DEFAULTS, **(revolver_config or {})}
    PCT_FMT = "0.00%"
    sheet_q = f"'{SHEET_NOTES}'"
    start = len(lines) + 2  # one blank gap after the text block

    ws.cell(row=start, column=1, value="Revolver Inputs").font = F_HEADER
    inputs = [
        ("Facility Total", "Facility_Total", float(cfg["facility_total"]), CURRENCY_FMT),
        ("LC Carve-out", "LC_CarveOut", float(cfg["lc_carve_out"]), CURRENCY_FMT),
        ("Currently Drawn (wk0 anchor)", "Currently_Drawn", float(cfg["current_drawn_balance"]), CURRENCY_FMT),
        ("Beginning Cash (wk1 anchor)", "Beginning_Cash_Wk1", float(cfg["beginning_cash"]), CURRENCY_FMT),
        ("Minimum Cash Target", "Min_Cash_Target", float(cfg["minimum_cash_target"]), CURRENCY_FMT),
        ("SOFR", "SOFR", float(cfg["sofr_rate"]), PCT_FMT),
        ("Spread", "Spread", float(cfg["spread"]), PCT_FMT),
    ]
    row = start + 1
    for label, name, value, fmt in inputs:
        ws.cell(row=row, column=1, value=label).font = F_HEADER
        cell = ws.cell(row=row, column=2, value=value)
        cell.fill = _INPUT_FILL
        cell.font = F_INPUT
        cell.number_format = fmt
        ws.parent.defined_names.add(DefinedName(name, attr_text=f"{sheet_q}!$B${row}"))
        row += 1

    # Computed cells (normal style, formulas off the named inputs).
    ws.cell(row=row, column=1, value="Annual Rate").font = F_HEADER
    ar_cell = ws.cell(row=row, column=2, value="=SOFR+Spread")
    ar_cell.number_format = PCT_FMT
    ar_cell.font = F_BASE
    ws.parent.defined_names.add(DefinedName("Annual_Rate", attr_text=f"{sheet_q}!$B${row}"))
    row += 1
    ws.cell(row=row, column=1, value="Max Capacity").font = F_HEADER
    mc_cell = ws.cell(row=row, column=2, value="=Facility_Total-LC_CarveOut")
    mc_cell.number_format = CURRENCY_FMT
    mc_cell.font = F_BASE
    ws.parent.defined_names.add(DefinedName("Max_Capacity", attr_text=f"{sheet_q}!$B${row}"))

    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 18


# ---------------------------------------------------------------------------
# Assembly + write
# ---------------------------------------------------------------------------


def _set_widths(ws: Worksheet, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_workbook(
    receipts: pd.DataFrame,
    disbursements: pd.DataFrame,
    customers: pd.DataFrame,
    vendors: pd.DataFrame,
    as_of_date: dt.date,
    refresh_ts: Optional[dt.datetime] = None,
    variance: Optional[pd.DataFrame] = None,
    combined: Optional[pd.DataFrame] = None,
    combined_disbursements: Optional[pd.DataFrame] = None,
    po_diagnostics: Optional[dict] = None,
    payroll: Optional[pd.DataFrame] = None,
    debt_principal: Optional[pd.DataFrame] = None,
    debt_interest: Optional[pd.DataFrame] = None,
    revolver_config: Optional[dict] = None,
    forecast_vs_actual: Optional[pd.DataFrame] = None,
    actuals: Optional[dict] = None,
    horizon_weeks: int = FORECAST_HORIZON_WEEKS,
) -> Workbook:
    """Build the ten-sheet forecast workbook (Phase 7f layout).

    Sheet order: 13-Week Forecast, AR by Customer, AP by Vendor, Payroll,
    Debt Service, Revolver, Variance, Actual vs Forecast, Actuals,
    Assumptions & Notes.

    The revolver plug is visible Excel math (formulas referencing the Assumptions
    named-range inputs); the Forecast references it. 'Actual vs Forecast' renders
    the forecast-accuracy grading; 'Actuals' is the CFO's weekly input tab
    (round-tripped to inputs/actuals.json by src.ingest_actuals). The Assumptions
    sheet is built last so its named ranges resolve against the final layout.
    """
    wb = Workbook()
    ws_fc = wb.active
    ws_fc.title = SHEET_FORECAST
    ws_ar = wb.create_sheet(SHEET_AR)
    ws_ap = wb.create_sheet(SHEET_AP)
    ws_pay = wb.create_sheet(SHEET_PAYROLL)
    ws_debt = wb.create_sheet(SHEET_DEBT)
    ws_rev = wb.create_sheet(SHEET_REVOLVER)
    ws_var = wb.create_sheet(SHEET_VARIANCE)
    ws_avf = wb.create_sheet(SHEET_AVF)
    ws_act = wb.create_sheet(SHEET_ACTUALS)
    ws_notes = wb.create_sheet(SHEET_NOTES)

    ar_long = combined if (combined is not None and not combined.empty) else receipts
    ap_long = (
        combined_disbursements
        if (combined_disbursements is not None and not combined_disbursements.empty)
        else disbursements
    )
    n_cust = _build_ar_by_customer_sheet(ws_ar, ar_long, customers, horizon_weeks)
    n_vend = _build_ap_by_vendor_sheet(ws_ap, ap_long, vendors, horizon_weeks)
    _build_payroll_sheet(ws_pay, payroll, horizon_weeks)
    debt_total_col = _build_debt_service_sheet(ws_debt, debt_principal, debt_interest, horizon_weeks)
    _build_forecast_sheet(
        ws_fc, as_of_date, horizon_weeks, n_cust, n_vend, debt_total_col=debt_total_col,
    )
    _build_revolver_sheet(ws_rev, as_of_date, horizon_weeks)
    _build_variance_sheet(ws_var, variance, horizon_weeks)
    _build_actual_vs_forecast_sheet(ws_avf, forecast_vs_actual)
    _build_actuals_sheet(ws_act, actuals, as_of_date, horizon_weeks)
    _build_assumptions_sheet(ws_notes, as_of_date, refresh_ts, po_diagnostics, revolver_config)

    return wb


def write_workbook(wb: Workbook, path) -> Path:
    """Save the workbook, creating the parent directory if needed."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def run(
    as_of_date: Optional[dt.date] = None,
    output_path: Optional[Path] = None,
) -> Path:
    """Entrypoint: load bucketed tables, build the workbook, write the .xlsx."""
    logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
    if as_of_date is None:
        as_of_date = dt.date.today()
    if output_path is None:
        output_path = DATA_DIR / OUTPUT_FILENAME

    data = load_inputs()
    refresh_ts = latest_refresh_timestamp(ONEDRIVE_DATA_PATH)
    logger.info(
        "Loaded %d receipt rows, %d disbursement rows; as_of=%s, refreshed=%s",
        len(data["receipts"]), len(data["disbursements"]),
        as_of_date.isoformat(),
        refresh_ts.isoformat(sep=" ", timespec="seconds") if refresh_ts else "unknown",
    )

    po_diagnostics = po_liabilities_diagnostics(data["po_payments"])
    revolver_config = load_revolver_config()
    actuals = load_actuals()
    wb = build_workbook(
        data["receipts"], data["disbursements"],
        data["customers"], data["vendors"],
        as_of_date, refresh_ts,
        variance=data["variance"], combined=data["combined"],
        combined_disbursements=data["combined_disbursements"],
        po_diagnostics=po_diagnostics,
        payroll=data["payroll"],
        debt_principal=data["debt_principal"],
        debt_interest=data["debt_interest"],
        revolver_config=revolver_config,
        forecast_vs_actual=data["forecast_vs_actual"],
        actuals=actuals,
    )
    saved = write_workbook(wb, output_path)
    logger.info("Wrote workbook to %s", saved)
    return saved


if __name__ == "__main__":
    run()
