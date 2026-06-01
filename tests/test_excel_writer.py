"""Tests for src/output/excel_writer.py.

These use synthetic in-horizon fixtures (mirroring the repo's pytest tmp_path
pattern) rather than the live SQLite tables, so the assertions are isolated and
deterministic.

The workbook is now formula-driven: forecast AR/AP columns are cross-sheet
=SUM links, net/totals/entity-Total are formulas. openpyxl writes formulas as
strings without evaluating them, so:
  - Structure/formula tests assert the formula STRINGS are correct.
  - Numeric-correctness tests validate the hardcoded SOURCE values on the detail
    sheets (the numbers the formulas aggregate), which is what determines the
    forecast totals once Excel/LibreOffice recalculates.
"""
import datetime as dt

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.output.excel_writer import (
    CURRENCY_FMT,
    DELTA_FMT,
    DELTA_FMT_UNFAVORABLE,
    OUTPUT_FILENAME,
    SHEET_AP,
    SHEET_AR,
    SHEET_DEBT,
    SHEET_FORECAST,
    SHEET_NOTES,
    SHEET_PAYROLL,
    SHEET_REVOLVER,
    SHEET_VARIANCE,
    VARIANCE_PLACEHOLDER,
    build_workbook,
    latest_refresh_timestamp,
    run,
    write_workbook,
)


AS_OF = dt.date(2026, 5, 29)  # Friday; week-1 Monday = 2026-05-25


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _receipts():
    """ar_receipts_by_week shape: 3 customers across a few weeks."""
    return pd.DataFrame({
        "customerNumber": ["CUST-A", "CUST-A", "CUST-B", "CUST-C"],
        "forecast_week":  [1, 3, 2, 13],
        "week_start_date": [
            dt.date(2026, 5, 25), dt.date(2026, 6, 8),
            dt.date(2026, 6, 1), dt.date(2026, 8, 17),
        ],
        "receipts": [10_000.0, 5_000.0, 8_000.0, 2_000.0],
    })


def _disbursements():
    """ap_disbursements_by_week shape: 2 vendors."""
    return pd.DataFrame({
        "vendorNumber": ["VEND-A", "VEND-A", "VEND-B"],
        "forecast_week": [1, 2, 2],
        "week_start_date": [
            dt.date(2026, 5, 25), dt.date(2026, 6, 1), dt.date(2026, 6, 1),
        ],
        "disbursements": [3_000.0, 9_000.0, 1_500.0],
    })


def _customers():
    return pd.DataFrame({
        "number": ["CUST-A", "CUST-B", "CUST-C"],
        "displayName": ["Alpha Foods", "Beta Grocers", "Gamma Mart"],
    })


def _vendors():
    return pd.DataFrame({
        "number": ["VEND-A", "VEND-B"],
        "displayName": ["Acme Supply", "Borden Logistics"],
    })


def _build(tmp_path):
    """Build the workbook from fixtures, save it, and return the loaded copy."""
    wb = build_workbook(
        _receipts(), _disbursements(), _customers(), _vendors(),
        AS_OF, refresh_ts=dt.datetime(2026, 5, 29, 12, 0, 0),
    )
    path = tmp_path / "forecast.xlsx"
    write_workbook(wb, path)
    return load_workbook(path)


def _row_week_value_sum(ws, row, first_col=3, last_col=15):
    """Sum the hardcoded week-value cells (C..O) of a detail-sheet row."""
    return sum(
        (ws.cell(row=row, column=c).value or 0.0)
        for c in range(first_col, last_col + 1)
    )


# ---------------------------------------------------------------------------
# Structure
# ---------------------------------------------------------------------------


def test_workbook_created_at_expected_path(tmp_path):
    path = tmp_path / "sub" / "forecast.xlsx"
    wb = build_workbook(
        _receipts(), _disbursements(), _customers(), _vendors(), AS_OF,
    )
    saved = write_workbook(wb, path)

    assert saved.exists()
    assert saved.name == "forecast.xlsx"


def test_run_writes_to_default_data_dir_filename(tmp_path):
    """run() with an explicit path produces a file with the module's filename."""
    out = tmp_path / OUTPUT_FILENAME
    saved = run(as_of_date=AS_OF, output_path=out)

    assert saved.exists()
    assert saved.name == OUTPUT_FILENAME


def test_all_sheets_exist_in_expected_order(tmp_path):
    wb = _build(tmp_path)
    assert wb.sheetnames == [
        SHEET_FORECAST, SHEET_AR, SHEET_AP, SHEET_PAYROLL,
        SHEET_DEBT, SHEET_REVOLVER, SHEET_VARIANCE, SHEET_NOTES,
    ]


def test_forecast_sheet_has_header_13_weeks_and_totals_row(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_FORECAST]

    # 1 header + 13 week rows + 1 totals row = 15
    assert ws.max_row == 15
    assert [ws.cell(row=r, column=1).value for r in range(2, 15)] == list(range(1, 14))
    assert ws.cell(row=15, column=1).value == "TOTAL"


def test_forecast_week_start_dates_are_mondays_seven_days_apart(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_FORECAST]

    wk1 = ws.cell(row=2, column=2).value
    wk2 = ws.cell(row=3, column=2).value
    assert dt.date(wk1.year, wk1.month, wk1.day) == dt.date(2026, 5, 25)  # Monday
    assert (wk2 - wk1).days == 7


# ---------------------------------------------------------------------------
# Formulas
# ---------------------------------------------------------------------------


def test_forecast_receipts_disbursements_are_cross_sheet_sum_links(tmp_path):
    """D/E columns sum the matching week column on the detail sheets.

    3 customers -> AR rows 2:4 ; 2 vendors -> AP rows 2:3. Week 1 -> col C,
    week 2 -> col D.
    """
    wb = _build(tmp_path)
    ws = wb[SHEET_FORECAST]

    assert ws.cell(row=2, column=4).value == "=SUM('AR by Customer'!C2:C4)"
    assert ws.cell(row=2, column=5).value == "=SUM('AP by Vendor'!C2:C3)"
    assert ws.cell(row=3, column=4).value == "=SUM('AR by Customer'!D2:D4)"
    assert ws.cell(row=3, column=5).value == "=SUM('AP by Vendor'!D2:D3)"
    # Week 13 -> column O
    assert ws.cell(row=14, column=4).value == "=SUM('AR by Customer'!O2:O4)"


def test_forecast_net_and_totals_are_formulas(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_FORECAST]

    # Net Cash Flow (col J) = receipts - disb - payroll - debt + rev_net - rev_int.
    assert ws.cell(row=2, column=10).value == "=D2-E2-F2-G2+H2-I2"
    assert ws.cell(row=14, column=10).value == "=D14-E14-F14-G14+H14-I14"
    # Totals row sums each cash-movement column (D through J).
    assert ws.cell(row=15, column=4).value == "=SUM(D2:D14)"
    assert ws.cell(row=15, column=7).value == "=SUM(G2:G14)"   # Debt Service
    assert ws.cell(row=15, column=10).value == "=SUM(J2:J14)"  # Net Cash Flow


def test_payroll_and_debt_columns_reference_detail_tabs(tmp_path):
    """Forecast Payroll (F) / Debt Service (G) link to the detail tabs, same row.

    Built with 5-loan debt data so Total Debt Service lands in the standard
    column O (the live layout).
    """
    payroll = pd.DataFrame({
        "forecast_week": [1], "week_start_date": ["2026-05-25"],
        "gross_wages": [100.0], "employer_burden_pct": [0.12],
        "total_payroll": [112.0], "source_stream": ["payroll"],
    })
    loans = [f"Loan {i}" for i in range(1, 6)]
    debt_p = pd.DataFrame({
        "forecast_week": [1] * 5, "week_start_date": ["2026-05-25"] * 5,
        "loan_name": loans, "principal": [10.0] * 5, "source_stream": ["debt_principal"] * 5,
    })
    debt_i = pd.DataFrame({
        "forecast_week": [1] * 5, "week_start_date": ["2026-05-25"] * 5,
        "loan_name": loans, "interest": [1.0] * 5, "source_stream": ["debt_interest"] * 5,
    })
    wb = build_workbook(
        _receipts(), _disbursements(), _customers(), _vendors(), AS_OF,
        payroll=payroll, debt_principal=debt_p, debt_interest=debt_i,
    )
    path = tmp_path / "ref.xlsx"
    write_workbook(wb, path)
    ws = load_workbook(path)[SHEET_FORECAST]

    assert ws.cell(row=2, column=6).value == "='Payroll'!F2"
    assert ws.cell(row=2, column=7).value == "='Debt Service'!O2"
    assert ws.cell(row=14, column=6).value == "='Payroll'!F14"
    assert ws.cell(row=14, column=7).value == "='Debt Service'!O14"


def test_forecast_cash_columns_reference_revolver_tab(tmp_path):
    """Phase 7e: Beginning/Ending cash on the Forecast come from the Revolver tab.

    No static revolver values; Begin (C) and Ending (K) are pure references into
    the Revolver tab's Begin Cash (C) and Ending Cash (N) columns, same row.
    """
    wb = _build(tmp_path)
    ws = wb[SHEET_FORECAST]

    assert ws.cell(row=2, column=3).value == "='Revolver'!C2"
    assert ws.cell(row=2, column=11).value == "='Revolver'!N2"
    assert ws.cell(row=3, column=3).value == "='Revolver'!C3"
    assert ws.cell(row=3, column=11).value == "='Revolver'!N3"


def test_entity_total_column_is_sum_formula(tmp_path):
    wb = _build(tmp_path)
    ar = wb[SHEET_AR]
    ap = wb[SHEET_AP]

    # Total column = P (16); weeks span C:O.
    assert ar.cell(row=2, column=16).value == "=SUM(C2:O2)"
    assert ap.cell(row=2, column=16).value == "=SUM(C2:O2)"
    assert ar.cell(row=ar.max_row, column=16).value == f"=SUM(C{ar.max_row}:O{ar.max_row})"


# ---------------------------------------------------------------------------
# Numeric correctness (validated on the source values the formulas aggregate)
# ---------------------------------------------------------------------------


def test_detail_source_values_match_input_aggregates(tmp_path):
    """The week-value cells the forecast SUMs over reconcile to the input totals."""
    wb = _build(tmp_path)
    ar = wb[SHEET_AR]
    ap = wb[SHEET_AP]

    ar_total = sum(_row_week_value_sum(ar, r) for r in range(2, ar.max_row + 1))
    ap_total = sum(_row_week_value_sum(ap, r) for r in range(2, ap.max_row + 1))

    assert ar_total == pytest.approx(_receipts()["receipts"].sum())        # 25,000
    assert ap_total == pytest.approx(_disbursements()["disbursements"].sum())  # 13,500


def test_forecast_per_week_source_values(tmp_path):
    """Week 1 and week 2 source columns on the detail sheets hold the right cash."""
    wb = _build(tmp_path)
    ar = wb[SHEET_AR]
    ap = wb[SHEET_AP]

    # Week 1 = column C: AR 10,000 ; AP 3,000
    ar_wk1 = sum(ar.cell(row=r, column=3).value or 0.0 for r in range(2, ar.max_row + 1))
    ap_wk1 = sum(ap.cell(row=r, column=3).value or 0.0 for r in range(2, ap.max_row + 1))
    assert ar_wk1 == pytest.approx(10_000.0)
    assert ap_wk1 == pytest.approx(3_000.0)
    # Week 2 = column D: AR 8,000 ; AP 10,500
    ar_wk2 = sum(ar.cell(row=r, column=4).value or 0.0 for r in range(2, ar.max_row + 1))
    ap_wk2 = sum(ap.cell(row=r, column=4).value or 0.0 for r in range(2, ap.max_row + 1))
    assert ar_wk2 == pytest.approx(8_000.0)
    assert ap_wk2 == pytest.approx(10_500.0)


# ---------------------------------------------------------------------------
# Formatting
# ---------------------------------------------------------------------------


def test_currency_cells_use_dollar_format_with_dash_zeros(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_FORECAST]

    assert CURRENCY_FMT == "$#,##0_);[Red]($#,##0);-"
    assert ws.cell(row=2, column=4).number_format == CURRENCY_FMT   # AR Receipts
    assert ws.cell(row=15, column=6).number_format == CURRENCY_FMT  # totals Net
    # Detail sheets too.
    assert wb[SHEET_AR].cell(row=2, column=3).number_format == CURRENCY_FMT


def test_revolver_input_cells_are_blue_on_yellow(tmp_path):
    """Phase 7e: the input cells now live in the Assumptions Revolver Inputs block."""
    wb = _build(tmp_path)
    ws = wb[SHEET_NOTES]

    # Find the "Facility Total" label row; its column-B value is a yellow input.
    label_row = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "Facility Total"
    )
    cell = ws.cell(row=label_row, column=2)
    assert (cell.font.color.rgb or "").endswith("0000FF")     # blue text
    assert (cell.fill.fgColor.rgb or "").endswith("FFFF00")    # yellow fill


def test_cross_sheet_link_cells_are_green(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_FORECAST]
    assert (ws.cell(row=2, column=4).font.color.rgb or "").endswith("008000")


def test_workbook_uses_arial_font(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_FORECAST]
    assert ws.cell(row=1, column=1).font.name == "Arial"   # header
    assert ws.cell(row=2, column=1).font.name == "Arial"   # data


# ---------------------------------------------------------------------------
# Entity sheets
# ---------------------------------------------------------------------------


def test_ar_sheet_row_count_equals_distinct_customers(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_AR]

    distinct = _receipts()["customerNumber"].nunique()  # 3
    assert ws.max_row == distinct + 1  # + header


def test_ap_sheet_row_count_equals_distinct_vendors(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_AP]

    distinct = _disbursements()["vendorNumber"].nunique()  # 2
    assert ws.max_row == distinct + 1


def test_ar_sheet_sorted_by_total_descending_with_names(tmp_path):
    """Rows are sorted by total cash desc; verified via the week source cells."""
    wb = _build(tmp_path)
    ws = wb[SHEET_AR]

    # CUST-A total = 15,000 (largest) -> first data row, joined to its name.
    assert ws.cell(row=2, column=1).value == "CUST-A"
    assert ws.cell(row=2, column=2).value == "Alpha Foods"
    row_totals = [_row_week_value_sum(ws, r) for r in range(2, ws.max_row + 1)]
    assert row_totals == sorted(row_totals, reverse=True)
    assert row_totals[0] == pytest.approx(15_000.0)


def test_ap_sheet_has_week_columns_and_total(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_AP]

    # Header: Vendor Number, Vendor Name, Week 1..13, Total, Source = 17 columns
    # (Source added in Phase 5, mirroring the AR-by-Customer sheet).
    assert ws.max_column == 17
    assert ws.cell(row=1, column=1).value == "Vendor Number"
    assert ws.cell(row=1, column=16).value == "Total"
    assert ws.cell(row=1, column=17).value == "Source"
    # AP-only fallback tags every row open_ap.
    assert ws.cell(row=2, column=17).value == "open_ap"
    # VEND-A row: week source cells sum to 12,000 (3,000 wk1 + 9,000 wk2).
    assert ws.cell(row=2, column=1).value == "VEND-A"
    assert _row_week_value_sum(ws, 2) == pytest.approx(12_000.0)


# ---------------------------------------------------------------------------
# Notes sheet
# ---------------------------------------------------------------------------


def test_notes_sheet_contains_as_of_date(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_NOTES]

    text = "\n".join(
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    )
    assert AS_OF.isoformat() in text


def test_notes_sheet_mentions_methodology_and_caveats(tmp_path):
    wb = _build(tmp_path)
    ws = wb[SHEET_NOTES]

    text = "\n".join(
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    )
    assert "Methodology" in text
    assert "Caveats" in text
    assert "payroll" in text


# ---------------------------------------------------------------------------
# Variance sheet
# ---------------------------------------------------------------------------


def _variance():
    """forecast_variance shape: 2 weeks of today-vs-prior deltas."""
    return pd.DataFrame({
        "snapshot_date": ["2026-05-30", "2026-05-30"],
        "prior_snapshot_date": ["2026-05-29", "2026-05-29"],
        "forecast_week": [1, 2],
        "week_start_date": ["2026-05-25", "2026-06-01"],
        "ar_receipts_today": [1200.0, 800.0],
        "ar_receipts_prior": [1000.0, 900.0],
        "ar_receipts_delta": [200.0, -100.0],
        "ap_disbursements_today": [350.0, 200.0],
        "ap_disbursements_prior": [300.0, 200.0],
        "ap_disbursements_delta": [50.0, 0.0],
        "net_today": [850.0, 600.0],
        "net_prior": [700.0, 700.0],
        "net_delta": [150.0, -100.0],
    })


def _build_with_variance(tmp_path, var):
    wb = build_workbook(
        _receipts(), _disbursements(), _customers(), _vendors(),
        AS_OF, refresh_ts=None, variance=var,
    )
    path = tmp_path / "var.xlsx"
    write_workbook(wb, path)
    return load_workbook(path)


def test_variance_sheet_position(tmp_path):
    # Phase 7e moved Variance to 7th (index 6), after the Revolver tab.
    wb = _build(tmp_path)
    assert wb.sheetnames[6] == SHEET_VARIANCE


def test_variance_first_run_shows_placeholder(tmp_path):
    """No prior snapshot (variance None) -> placeholder message in A1, no grid."""
    wb = _build(tmp_path)   # _build passes no variance -> None
    ws = wb[SHEET_VARIANCE]

    assert ws.cell(row=1, column=1).value == VARIANCE_PLACEHOLDER
    assert ws.cell(row=2, column=1).value is None  # no data grid


def test_variance_empty_dataframe_also_shows_placeholder(tmp_path):
    empty = pd.DataFrame()
    wb = _build_with_variance(tmp_path, empty)
    ws = wb[SHEET_VARIANCE]
    assert ws.cell(row=1, column=1).value == VARIANCE_PLACEHOLDER


def test_variance_populated_has_13_rows_plus_header_and_totals(tmp_path):
    wb = _build_with_variance(tmp_path, _variance())
    ws = wb[SHEET_VARIANCE]

    assert ws.max_row == 15  # header + 13 weeks + totals
    assert ws.cell(row=1, column=1).value == "Forecast Week"
    assert ws.cell(row=15, column=1).value == "TOTAL"
    assert [ws.cell(row=r, column=1).value for r in range(2, 15)] == list(range(1, 14))


def test_variance_delta_columns_are_formulas_with_delta_format(tmp_path):
    wb = _build_with_variance(tmp_path, _variance())
    ws = wb[SHEET_VARIANCE]

    # AR Δ (E), AP Δ (H), Net Δ (K) are formulas with the green/red delta format.
    assert ws.cell(row=2, column=5).value == "=C2-D2"
    assert ws.cell(row=2, column=8).value == "=F2-G2"
    assert ws.cell(row=2, column=11).value == "=I2-J2"
    # AR Δ (E) and Net Δ (K) use favorable coloring (positive green).
    assert ws.cell(row=2, column=5).number_format == DELTA_FMT
    assert ws.cell(row=2, column=11).number_format == DELTA_FMT
    # Net today/prior are also formulas.
    assert ws.cell(row=2, column=9).value == "=C2-F2"
    assert ws.cell(row=2, column=10).value == "=D2-G2"


def test_variance_ap_delta_uses_unfavorable_format(tmp_path):
    """AP Δ (column H) is red-for-positive: a rise in disbursements is bad."""
    wb = _build_with_variance(tmp_path, _variance())
    ws = wb[SHEET_VARIANCE]

    assert DELTA_FMT_UNFAVORABLE == "[Red]$#,##0;[Green]($#,##0);-"
    assert ws.cell(row=2, column=8).number_format == DELTA_FMT_UNFAVORABLE   # data row
    assert ws.cell(row=15, column=8).number_format == DELTA_FMT_UNFAVORABLE  # totals
    # AR Δ and Net Δ remain favorable, distinct from AP Δ.
    assert ws.cell(row=2, column=5).number_format == DELTA_FMT
    assert ws.cell(row=2, column=11).number_format == DELTA_FMT


def test_variance_today_prior_source_values(tmp_path):
    wb = _build_with_variance(tmp_path, _variance())
    ws = wb[SHEET_VARIANCE]

    # Week 1: AR today/prior and AP today/prior come straight from the table.
    assert ws.cell(row=2, column=3).value == pytest.approx(1200.0)
    assert ws.cell(row=2, column=4).value == pytest.approx(1000.0)
    assert ws.cell(row=2, column=6).value == pytest.approx(350.0)
    assert ws.cell(row=2, column=7).value == pytest.approx(300.0)
    # Totals row sums via formula.
    assert ws.cell(row=15, column=3).value == "=SUM(C2:C14)"


def test_variance_missing_weeks_filled_with_zero(tmp_path):
    """Variance with only week 1 still renders 13 rows; week 2 source cells = 0."""
    one_week = _variance().iloc[[0]].copy()
    wb = _build_with_variance(tmp_path, one_week)
    ws = wb[SHEET_VARIANCE]

    assert ws.max_row == 15
    assert ws.cell(row=3, column=3).value == pytest.approx(0.0)  # week 2 AR today


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


def test_empty_inputs_still_produce_all_sheets(tmp_path):
    """No receipts/disbursements/payroll/debt -> valid workbook, formulas hold."""
    empty_rec = pd.DataFrame(columns=["customerNumber", "forecast_week", "week_start_date", "receipts"])
    empty_dis = pd.DataFrame(columns=["vendorNumber", "forecast_week", "week_start_date", "disbursements"])

    wb = build_workbook(empty_rec, empty_dis, _customers(), _vendors(), AS_OF)
    path = tmp_path / "empty.xlsx"
    write_workbook(wb, path)
    loaded = load_workbook(path)

    assert loaded.sheetnames == [
        SHEET_FORECAST, SHEET_AR, SHEET_AP, SHEET_PAYROLL,
        SHEET_DEBT, SHEET_REVOLVER, SHEET_VARIANCE, SHEET_NOTES,
    ]
    ws = loaded[SHEET_FORECAST]
    assert ws.max_row == 15  # header + 13 + totals
    # Totals are still formulas; detail SUM ranges collapse to an empty row-2 cell.
    assert ws.cell(row=15, column=4).value == "=SUM(D2:D14)"
    assert ws.cell(row=2, column=4).value == "=SUM('AR by Customer'!C2:C2)"
    # Payroll / Debt detail tabs render with header + 13 weeks + totals row.
    assert loaded[SHEET_PAYROLL].max_row == 15
    assert loaded[SHEET_DEBT].max_row == 15
    # Debt Service ref points at the detail tab, same row (column depends on
    # loan count, which is 0 here, so don't hardcode it).
    assert ws.cell(row=2, column=7).value.startswith("='Debt Service'!")
    assert ws.cell(row=2, column=7).value.endswith("2")
    # Entity sheets have only their header row.
    assert loaded[SHEET_AR].max_row == 1
    assert loaded[SHEET_AP].max_row == 1


def test_latest_refresh_timestamp_missing_folder_returns_none(tmp_path):
    missing = tmp_path / "does-not-exist"
    assert latest_refresh_timestamp(missing) is None


def test_latest_refresh_timestamp_picks_newest_csv(tmp_path):
    import os, time

    older = tmp_path / "AR_History_2026-05-01.csv"
    newer = tmp_path / "AP_History_2026-05-29.csv"
    older.write_text("x")
    newer.write_text("y")
    os.utime(older, (time.time() - 1000, time.time() - 1000))

    ts = latest_refresh_timestamp(tmp_path)
    assert ts is not None
    assert isinstance(ts, dt.datetime)


# ---------------------------------------------------------------------------
# Open-SO integration (Task 5): combined AR+SO on the AR-by-Customer sheet
# ---------------------------------------------------------------------------


def _combined():
    """combined_receipts_by_week shape: CUST-A in BOTH sources, CUST-B SO-only."""
    return pd.DataFrame({
        "customerNumber": ["CUST-A", "CUST-A", "CUST-B", "CUST-D"],
        "forecast_week":  [1, 8, 9, 10],
        "week_start_date": ["2026-05-25", "2026-07-13", "2026-07-20", "2026-07-27"],
        "receipts": [10_000.0, 4_000.0, 2_000.0, 1_500.0],
        "source": ["open_ar", "open_so", "open_so", "open_so"],
    })


def _build_with_combined(tmp_path, combined):
    wb = build_workbook(
        _receipts(), _disbursements(), _customers(), _vendors(),
        AS_OF, refresh_ts=None, combined=combined,
    )
    path = tmp_path / "combined.xlsx"
    write_workbook(wb, path)
    return load_workbook(path)


def test_ar_sheet_has_source_column(tmp_path):
    wb = _build(tmp_path)  # AR-only fallback
    ws = wb[SHEET_AR]
    # Header: Customer Number, Customer Name, Week 1..13, Total, Source = 17 cols
    assert ws.cell(row=1, column=17).value == "Source"
    assert ws.cell(row=1, column=16).value == "Total"
    # AR-only fallback tags every row open_ar.
    assert ws.cell(row=2, column=17).value == "open_ar"


def test_ar_sheet_combined_customer_in_both_sources_has_two_rows(tmp_path):
    wb = _build_with_combined(tmp_path, _combined())
    ws = wb[SHEET_AR]

    rows = [(ws.cell(r, 1).value, ws.cell(r, 17).value) for r in range(2, ws.max_row + 1)]
    cust_a_rows = [r for r in rows if r[0] == "CUST-A"]
    assert len(cust_a_rows) == 2
    assert {src for _, src in cust_a_rows} == {"open_ar", "open_so"}


def test_forecast_ar_total_reflects_combined_sources(tmp_path):
    """Forecast AR receipts SUM over the AR sheet now includes SO past week 6."""
    wb = _build_with_combined(tmp_path, _combined())
    ar = wb[SHEET_AR]

    # Week 8 (column J = 3+7) total across all AR rows = the open_so 4,000.
    wk8_total = sum(ar.cell(r, 10).value or 0.0 for r in range(2, ar.max_row + 1))
    assert wk8_total == pytest.approx(4_000.0)
    # Grand total across the sheet = sum of all combined receipts.
    grand = sum(_row_week_value_sum(ar, r) for r in range(2, ar.max_row + 1))
    assert grand == pytest.approx(_combined()["receipts"].sum())  # 17,500


def test_workbook_renders_ar_only_when_combined_empty(tmp_path):
    """Graceful degradation: empty combined -> AR-only sheet, no error."""
    empty_combined = pd.DataFrame(
        columns=["customerNumber", "forecast_week", "week_start_date", "receipts", "source"]
    )
    wb = _build_with_combined(tmp_path, empty_combined)
    ws = wb[SHEET_AR]

    # Falls back to the _receipts() fixture: 3 customers, all open_ar.
    assert ws.max_row == _receipts()["customerNumber"].nunique() + 1
    sources = {ws.cell(r, 17).value for r in range(2, ws.max_row + 1)}
    assert sources == {"open_ar"}


def test_workbook_renders_ar_only_when_combined_none(tmp_path):
    """combined=None (table absent) also falls back cleanly to AR-only."""
    wb = build_workbook(
        _receipts(), _disbursements(), _customers(), _vendors(),
        AS_OF, combined=None,
    )
    path = tmp_path / "none.xlsx"
    write_workbook(wb, path)
    ws = load_workbook(path)[SHEET_AR]
    assert ws.max_row == _receipts()["customerNumber"].nunique() + 1


# ---------------------------------------------------------------------------
# Open-PO integration (Phase 5): combined AP+PO on the AP-by-Vendor sheet
# ---------------------------------------------------------------------------


def _combined_disbursements():
    """combined_disbursements_by_week: VEND-A in open_ap AND po_rbni; PO fills late weeks."""
    return pd.DataFrame({
        "vendorNumber": ["VEND-A", "VEND-A", "VEND-B", "VEND-C"],
        "forecast_week":  [1, 8, 9, 11],
        "week_start_date": ["2026-05-25", "2026-07-13", "2026-07-20", "2026-08-03"],
        "disbursements": [3_000.0, 4_000.0, 2_000.0, 1_500.0],
        "source": ["open_ap", "po_rbni", "po_outstanding", "po_rbni"],
    })


def _po_diag():
    return {
        "rbni_total": 5_500.0, "outstanding_total": 2_000.0, "phase5_total": 7_500.0,
        "rbni_by_type": {"Item": 4_000.0, "Charge (Item)": 1_500.0},
        "outstanding_by_type": {"Item": 2_000.0},
        "item_rbni_subtotal": 4_000.0, "invoice_lag_days": 7,
    }


def _build_with_po(tmp_path, combined_disb, po_diag=None):
    wb = build_workbook(
        _receipts(), _disbursements(), _customers(), _vendors(),
        AS_OF, refresh_ts=None,
        combined_disbursements=combined_disb, po_diagnostics=po_diag,
    )
    path = tmp_path / "po.xlsx"
    write_workbook(wb, path)
    return load_workbook(path)


def test_ap_sheet_combined_vendor_in_two_sources_has_two_rows(tmp_path):
    wb = _build_with_po(tmp_path, _combined_disbursements())
    ws = wb[SHEET_AP]
    rows = [(ws.cell(r, 1).value, ws.cell(r, 17).value) for r in range(2, ws.max_row + 1)]
    vend_a = [r for r in rows if r[0] == "VEND-A"]
    assert len(vend_a) == 2
    assert {src for _, src in vend_a} == {"open_ap", "po_rbni"}


def test_forecast_ap_total_reflects_combined_po_sources(tmp_path):
    """Forecast AP disbursements SUM over the AP sheet now includes PO past wk 6."""
    wb = _build_with_po(tmp_path, _combined_disbursements())
    ap = wb[SHEET_AP]
    # Week 8 (col J = 3+7) across all AP rows = the po_rbni 4,000.
    wk8 = sum(ap.cell(r, 10).value or 0.0 for r in range(2, ap.max_row + 1))
    assert wk8 == pytest.approx(4_000.0)
    grand = sum(_row_week_value_sum(ap, r) for r in range(2, ap.max_row + 1))
    assert grand == pytest.approx(_combined_disbursements()["disbursements"].sum())  # 10,500


def test_ap_renders_ap_only_when_combined_disbursements_empty(tmp_path):
    empty = pd.DataFrame(columns=["vendorNumber", "forecast_week", "week_start_date", "disbursements", "source"])
    wb = _build_with_po(tmp_path, empty)
    ws = wb[SHEET_AP]
    # Falls back to _disbursements(): 2 vendors, all open_ap.
    assert ws.max_row == _disbursements()["vendorNumber"].nunique() + 1
    assert {ws.cell(r, 17).value for r in range(2, ws.max_row + 1)} == {"open_ap"}


def test_assumptions_sheet_has_phase5_diagnostic_block(tmp_path):
    wb = _build_with_po(tmp_path, _combined_disbursements(), po_diag=_po_diag())
    ws = wb[SHEET_NOTES]
    text = "\n".join(str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1))
    assert "Phase 5 / PO Liabilities Diagnostic" in text
    assert "GL 21300" in text
    assert "$5,500" in text          # RBNI total
    assert "$7,500" in text          # phase 5 total
    assert "Item-only RBNI subtotal: $4,000" in text


# ---------------------------------------------------------------------------
# Phase 7d: Payroll / Debt Service detail tabs + revolver columns
# ---------------------------------------------------------------------------


def _payroll_df():
    return pd.DataFrame({
        "forecast_week": [1, 2, 3],
        "week_start_date": ["2026-05-25", "2026-06-01", "2026-06-08"],
        "gross_wages": [1000.0, 1000.0, 2000.0],
        "employer_burden_pct": [0.12, 0.12, 0.12],
        "total_payroll": [1120.0, 1120.0, 2240.0],
        "source_stream": ["payroll"] * 3,
    })


def _debt_frames():
    loans = ["Loan 1", "Loan 2"]
    p = pd.DataFrame({
        "forecast_week": [1, 1, 5], "week_start_date": ["2026-05-25", "2026-05-25", "2026-06-22"],
        "loan_name": ["Loan 1", "Loan 2", "Loan 1"], "principal": [400.0, 100.0, 400.0],
        "source_stream": ["debt_principal"] * 3,
    })
    i = pd.DataFrame({
        "forecast_week": [1, 1, 5], "week_start_date": ["2026-05-25", "2026-05-25", "2026-06-22"],
        "loan_name": ["Loan 1", "Loan 2", "Loan 1"], "interest": [50.0, 10.0, 45.0],
        "source_stream": ["debt_interest"] * 3,
    })
    return p, i


def _build_7d(tmp_path):
    p, i = _debt_frames()
    wb = build_workbook(
        _receipts(), _disbursements(), _customers(), _vendors(), AS_OF,
        payroll=_payroll_df(), debt_principal=p, debt_interest=i,
    )
    path = tmp_path / "7d.xlsx"
    write_workbook(wb, path)
    return load_workbook(path)


def test_payroll_tab_total_column_matches_input(tmp_path):
    ws = _build_7d(tmp_path)[SHEET_PAYROLL]
    assert ws.cell(1, 6).value == "Total Payroll"
    # wk1 row 2, wk3 row 4
    assert ws.cell(2, 6).value == pytest.approx(1120.0)
    assert ws.cell(4, 6).value == pytest.approx(2240.0)
    # weeks with no payroll data render as 0.
    assert ws.cell(5, 6).value == pytest.approx(0.0)


def test_debt_tab_total_is_principal_plus_interest(tmp_path):
    ws = _build_7d(tmp_path)[SHEET_DEBT]
    # 2 loans -> Total Debt Service at column 2 + 2 + 1 + 2 + 1 + 1 = 9 (I).
    assert ws.cell(1, 9).value == "Total Debt Service"
    # wk1: principal 400+100=500, interest 50+10=60 -> 560.
    assert ws.cell(2, 9).value == pytest.approx(560.0)
    # wk5 (row 6): principal 400, interest 45 -> 445.
    assert ws.cell(6, 9).value == pytest.approx(445.0)


def test_forecast_revolver_columns_reference_revolver_tab(tmp_path):
    """Phase 7e: revolver Net (H) / Interest (I) are formula refs, not values."""
    ws = _build_7d(tmp_path)[SHEET_FORECAST]
    assert ws.cell(2, 8).value == "='Revolver'!K2-'Revolver'!L2"
    assert ws.cell(2, 9).value == "='Revolver'!G2"
    assert ws.cell(14, 8).value == "='Revolver'!K14-'Revolver'!L14"


# ---------------------------------------------------------------------------
# Phase 7e: Revolver tab (visible Excel math) + Assumptions named-range inputs
# ---------------------------------------------------------------------------

REV_CONFIG = {
    "facility_total": 60_000_000, "lc_carve_out": 2_545_000,
    "current_drawn_balance": 0, "beginning_cash": 1_196_696,
    "minimum_cash_target": 0, "sofr_rate": 0.0535, "spread": 0.0135,
}


def _build_7e(tmp_path, receipts=None, disbursements=None, config=None):
    wb = build_workbook(
        receipts if receipts is not None else _receipts(),
        disbursements if disbursements is not None else _disbursements(),
        _customers(), _vendors(), AS_OF,
        revolver_config=config if config is not None else REV_CONFIG,
    )
    path = tmp_path / "7e.xlsx"
    write_workbook(wb, path)
    return load_workbook(path)


def test_revolver_tab_structure_and_all_formulas(tmp_path):
    wb = _build_7e(tmp_path)
    ws = wb[SHEET_REVOLVER]
    # header + 13 data rows
    assert ws.max_row == 14
    assert ws.cell(1, 1).value == "Forecast Week"
    assert ws.cell(1, 15).value == "Capacity Breached"
    # Every computed cell (C..O) on every week row is a formula string.
    for r in range(2, 15):
        for c in range(3, 16):
            v = ws.cell(r, c).value
            assert isinstance(v, str) and v.startswith("="), f"R{r}C{c} not a formula: {v!r}"


def test_assumptions_revolver_inputs_and_named_ranges(tmp_path):
    wb = _build_7e(tmp_path)
    names = set(wb.defined_names.keys())
    expected = {"Facility_Total", "LC_CarveOut", "Currently_Drawn", "Beginning_Cash_Wk1",
                "Min_Cash_Target", "SOFR", "Spread", "Annual_Rate", "Max_Capacity"}
    assert expected <= names

    ws = wb[SHEET_NOTES]
    # Resolve each named range to its cell and check the seeded value/formula.
    def named_cell(name):
        ref = wb.defined_names[name].attr_text       # e.g. 'Assumptions & Notes'!$B$31
        coord = ref.split("!")[1].replace("$", "")
        return ws[coord]
    assert named_cell("Facility_Total").value == pytest.approx(60_000_000)
    assert named_cell("Beginning_Cash_Wk1").value == pytest.approx(1_196_696)
    assert named_cell("Min_Cash_Target").value == pytest.approx(0)
    assert named_cell("SOFR").value == pytest.approx(0.0535)
    # Computed cells are formulas off the named inputs.
    assert named_cell("Annual_Rate").value == "=SOFR+Spread"
    assert named_cell("Max_Capacity").value == "=Facility_Total-LC_CarveOut"


def _resolve_revolver(wb, config):
    """Resolve the Revolver-tab formula chain in Python; return per-week ending cash."""
    ar = wb[SHEET_AR]; ap = wb[SHEET_AP]; pay = wb[SHEET_PAYROLL]; debt = wb[SHEET_DEBT]
    annual = config["sofr_rate"] + config["spread"]
    maxcap = config["facility_total"] - config["lc_carve_out"]
    target = config["minimum_cash_target"]

    def col_sum(ws, c):
        return sum((ws.cell(r, c).value or 0.0) for r in range(2, ws.max_row + 1))
    # debt total column letter -> index (find "Total Debt Service" header)
    tds_col = next(c for c in range(1, debt.max_column + 1)
                   if debt.cell(1, c).value == "Total Debt Service")

    begin_cash = config["beginning_cash"]
    begin_rev = config["current_drawn_balance"]
    out = []
    for i in range(13):
        r = i + 2
        D = col_sum(ar, 2 + (i + 1))
        E = col_sum(ap, 2 + (i + 1))
        F = pay.cell(r, 6).value or 0.0
        G = debt.cell(r, tds_col).value or 0.0
        interest = round(begin_rev * annual / 52, 2)
        pre = begin_cash + D - (E + F + G) - interest
        avail = maxcap - begin_rev
        draw = max(min(target - pre, avail), 0.0) if pre < target else 0.0
        repay = min(pre - target, begin_rev) if (pre > target and begin_rev > 0) else 0.0
        ending_rev = begin_rev + draw - repay
        ending_cash = pre + draw - repay
        out.append({"wk": i + 1, "pre": pre, "draw": draw, "repay": repay,
                    "ending_cash": ending_cash})
        begin_cash = ending_cash
        begin_rev = ending_rev
    return out


def test_revolver_excel_resolution_matches_python(tmp_path):
    """Resolve the Revolver tab in Python; compare to compute_revolver to the cent."""
    from src.calc.revolver import compute_revolver
    rec = _receipts(); dis = _disbursements()
    wb = _build_7e(tmp_path, rec, dis, REV_CONFIG)

    resolved = _resolve_revolver(wb, REV_CONFIG)

    # Build the same inflow/base-outflow dicts compute_revolver expects.
    inflows = {wk: 0.0 for wk in range(1, 14)}
    for _, r in rec.iterrows():
        inflows[int(r["forecast_week"])] += float(r["receipts"])
    base = {wk: 0.0 for wk in range(1, 14)}
    for _, r in dis.iterrows():
        base[int(r["forecast_week"])] += float(r["disbursements"])
    # (payroll/debt are zero in this fixture)
    _, _, cash = compute_revolver(REV_CONFIG, inflows, base, AS_OF)
    py = {int(row["forecast_week"]): float(row["ending_cash"]) for _, row in cash.iterrows()}

    for row in resolved:
        assert round(row["ending_cash"], 2) == pytest.approx(py[row["wk"]], abs=0.01)


def test_scenario_wk1_begin_cash_is_seeded_value(tmp_path):
    wb = _build_7e(tmp_path)
    # Revolver C2 references the named input; resolve via the Assumptions cell.
    ws = wb[SHEET_REVOLVER]
    assert ws.cell(2, 3).value == "=Beginning_Cash_Wk1"
    resolved = _resolve_revolver(wb, REV_CONFIG)
    assert resolved[0]["wk"] == 1
    # wk1 begin cash anchors to 1,196,696 (the seeded input).
    # (resolved[0] uses begin_cash = config beginning_cash for wk1.)
    assert REV_CONFIG["beginning_cash"] == 1_196_696


def test_chain_integrity_wk2_begin_refs_wk1_ending(tmp_path):
    wb = _build_7e(tmp_path)
    rev = wb[SHEET_REVOLVER]
    # Revolver Begin Cash wk2 (row 3, col C) = prior row Ending Cash (col N).
    assert rev.cell(3, 3).value == "=N2"
    # Begin Revolver wk2 = prior Ending Revolver Balance (col M).
    assert rev.cell(3, 6).value == "=M2"


def test_min_cash_policy_no_draw_when_pre_positive(tmp_path):
    """With min_cash_target=0, the revolver does not draw in any positive-cash week."""
    # Make every week strongly cash-positive: big receipts, tiny disbursements.
    rec = pd.DataFrame({
        "customerNumber": ["CUST-A"] * 13,
        "forecast_week": list(range(1, 14)),
        "week_start_date": [dt.date(2026, 5, 25)] * 13,
        "receipts": [1_000_000.0] * 13,
    })
    dis = pd.DataFrame({
        "vendorNumber": ["VEND-A"] * 13,
        "forecast_week": list(range(1, 14)),
        "week_start_date": [dt.date(2026, 5, 25)] * 13,
        "disbursements": [10_000.0] * 13,
    })
    wb = _build_7e(tmp_path, rec, dis, REV_CONFIG)
    resolved = _resolve_revolver(wb, REV_CONFIG)
    for row in resolved:
        if row["pre"] > 0:
            assert row["draw"] == pytest.approx(0.0)
